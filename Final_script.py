#%% Importing functions

import os
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

#%% Changing the directory to file with data

os.chdir(r'/Users/thijshanssen/Desktop/THESIS/Data/')

#%% Loading the data

""" Orignial copies of the data where no changes will be made """

OG_FL = pd.read_csv('OG_FL_Finaal.csv') 
OG_WA = pd.read_csv('OG_WA_Finaal.csv')

""" Data copies where changes wille be made """

FL = OG_FL.copy()
WA = OG_WA.copy()

#%% Deleting unnecessary columns

deleted = ['StartDate', 'EndDate', 'Status', 'IPAddress', 'Finished', 
           'RecordedDate', 'RecipientLastName', 'RecipientFirstName',
           'RecipientEmail', 'ExternalReference', 'LocationLatitude', 
           'LocationLongitude', 'DistributionChannel',
           'UserLanguage', 'psid', 'entrySigCheck', 'verification', 
           'expectedSignature', 'errorMessage', 'completeLink',
           'screenOutLink', 'overQuotaLink', 'invalidSignatureLink', 
           'customLink1', 'customLink2', 'customLink3']

FL = FL.drop(columns=deleted)
WA = WA.drop(columns=deleted)
del deleted

#%% Deleting the first two rows: these contain unnecessary information

FL = FL.iloc[2:] 
WA = WA.iloc[2:]
FL.reset_index(drop=True, inplace=True)
WA.reset_index(drop=True, inplace=True)

#%% Add new IDs

""" Creating new IDs for all instances. ID's from the Flanders dataset start
with a 1, signifying their origence and ID's from the Walloon dataset start
with a 2. After that initial number, the IDs show the row position of the 
instance in the original dataset. """

        

OG_FL.insert(0, 'ID', np.nan)
OG_WA.insert(0, 'ID', np.nan)
OG_FL.loc[2:, 'ID'] = [f'1,{i}' for i in range(1, len(OG_FL) - 1)]
OG_WA.loc[2:, 'ID'] = [f'2,{i}' for i in range(1, len(OG_WA) - 1)]

FL.insert(0, 'ID', [f'1,{i}' for i in range(1, len(FL) + 1)])
WA.insert(0, 'ID', [f'2,{i}' for i in range(1, len(WA) + 1)])

#%% Add region identifier

FL.insert(1, 'Region', 'Flanders')
WA.insert(1, 'Region', 'Wallonia')

#%% encoding postal codes

""" Ensuring uniform values """

FL['Vraag 21'] = pd.to_numeric(FL['Vraag 21'], errors='coerce').fillna(0)
FL['Vraag 21'] = FL['Vraag 21'].astype(int)

WA['Vraag 21'] = pd.to_numeric(WA['Vraag 21'], errors='coerce').fillna(0)
WA['Vraag 21'] = WA['Vraag 21'].astype(int)

def categorize_postal(value):
    value = int(value)
    if 1000 <= value <= 1299:
        return 'Brussels'
    elif 1300 <= value <= 1499:
        return 'Walloon Brabant'
    elif 1500 <= value <= 1999 or 3000 <= value <= 3499:
        return 'Flemish Brabant'
    elif 2000 <= value <= 2999:
        return 'Antwerp'
    elif 3500 <= value <= 3999:
        return 'Limburg'
    elif 4000 <= value <= 4999:
        return 'Liège'
    elif 5000 <= value <= 5999:
        return 'Namur'
    elif 6000 <= value <= 6599 or 7000 <= value <= 7999:
        return 'Hainaut'
    elif 6600 <= value <= 6999:
        return 'Luxembourg'
    elif 8000 <= value <= 8999:
        return 'West Flanders'
    elif 9000 <= value <= 9999:
        return 'East Flanders'
    else:
        return 'Unknown'
FL['Vraag 21'] = FL['Vraag 21'].apply(categorize_postal)
WA['Vraag 21'] = WA['Vraag 21'].apply(categorize_postal)

""" Renaming the column """

FL.rename(columns={'Vraag 21': 'Province'}, inplace=True)
WA.rename(columns={'Vraag 21': 'Province'}, inplace=True)

""" Correcting people's region """

FL.rename(columns={'Intro 1 ': 'Intro 1'}, inplace=True)

WA_filter = FL['Province'].isin(['Walloon Brabant', 'Liège', 'Namur', 
                                 'Hainaut', 'Luxembourg'])
WA_filtered_count = len(FL[WA_filter])
WA = pd.concat([WA, FL[WA_filter]])
WA.loc[WA['ID'].str.startswith('1'), 'Region'] = 'Wallonia'
FL = FL[~WA_filter]

FL_filter = WA['Province'].isin(['Flemish Brabant', 'Antwerp', 'Limburg', 
                                 'East Flanders', 'West Flanders'])
FL_filtered_count = len(WA[FL_filter])
FL = pd.concat([FL, WA[FL_filter]])
FL.loc[FL['ID'].str.startswith('2'), 'Region'] = 'Flanders'
WA = WA[~FL_filter]

print("Walloons in the Flemish dataset:", WA_filtered_count)
print("Flemish people in the Waloon dataset:", FL_filtered_count)

FL.reset_index(drop=True, inplace=True)
WA.reset_index(drop=True, inplace=True)

del FL_filter, FL_filtered_count, WA_filter, WA_filtered_count

#%% Creating a table with deletions based on planned inclusion factors

data = {'FL': [], 'WA': []} # new variable to make a list


""" Uniform values """ 

mapping = {
    'Niet akkoord om deel te nemen': 'No',
    'Non': 'No',
    'Nooit': 'No',
    'Jamais': 'No',
    'Nee': 'No',
    'Non, jamais': 'No'
}

WA['Intro 1'] = WA['Intro 1'].replace(mapping)
FL['Intro 1'] = FL['Intro 1'].replace(mapping)
WA['Q159'] = WA['Q159'].replace(mapping)
FL['Q159'] = FL['Q159'].replace(mapping)
WA['Inclusion 3'] = WA['Inclusion 3'].replace(mapping)
FL['Inclusion 3'] = FL['Inclusion 3'].replace(mapping)
WA['Q152'] = WA['Q152'].replace(mapping)
FL['Q152'] = FL['Q152'].replace(mapping)
WA['Inclusion 4'] = WA['Inclusion 4'].replace(mapping)
FL['Inclusion 4'] = FL['Inclusion 4'].replace(mapping)

""" Statistics on people that did not finish the survey """

FL['Progress'] = pd.to_numeric(FL['Progress'], errors='coerce')
WA['Progress'] = pd.to_numeric(WA['Progress'], errors='coerce')
FL_unfinished = len(FL[FL['Progress'] < 100])
WA_unfinished = len(WA[WA['Progress'] < 100])
data['FL'].append(FL_unfinished)
data['WA'].append(WA_unfinished)

""" People that speeded through the survey """
FL['Duration (in seconds)'] = pd.to_numeric(FL['Duration (in seconds)'], 
                                            errors='coerce')
WA['Duration (in seconds)'] = pd.to_numeric(WA['Duration (in seconds)'], 
                                            errors='coerce')
fl_speeders_count = len(FL[FL['Duration (in seconds)'] < 300])
wa_speeders_count = len(WA[WA['Duration (in seconds)'] < 300])
data['FL'].append(fl_speeders_count)
data['WA'].append(wa_speeders_count)

""" People who did not consent to participate in the survey """

fl_nietakkoord = len(FL[FL['Intro 1'] == 'No'])
wa_nietakkoord = len(WA[WA['Intro 1'] == 'No'])
data['FL'].append(fl_nietakkoord)
data['WA'].append(wa_nietakkoord)

""" People who never cook """

fl_nooit_count = len(FL[FL['Q159'] == 'No'])
wa_nooit_count = len(WA[WA['Q159'] == 'No'])
data['FL'].append(fl_nooit_count)
data['WA'].append(wa_nooit_count)

""" People who don't use cooking cream """

fl_nooit_count = len(FL[FL['Inclusion 3'] == 'No'])
wa_nooit_count = len(WA[WA['Inclusion 3'] == 'No'])
data['FL'].append(fl_nooit_count)
data['WA'].append(wa_nooit_count)

""" Non Belgian residents """
fl_nee_count = len(FL[FL['Q152'] == 'No'])
wa_nee_count = len(WA[WA['Q152'] == 'No'])
data['FL'].append(fl_nee_count)
data['WA'].append(wa_nee_count)

""" Don't keep cooking cream in fridge """

fl_nooit_count = len(FL[FL['Inclusion 4'] == 'No'])
wa_nooit_count = len(WA[WA['Inclusion 4'] == 'No'])
data['FL'].append(fl_nooit_count)
data['WA'].append(wa_nooit_count)

""" Printint the data table """
print(pd.DataFrame(data))

del data, fl_nee_count, fl_nietakkoord, fl_nooit_count, fl_speeders_count, 
FL_unfinished, wa_nietakkoord, wa_nooit_count, 
wa_speeders_count, WA_unfinished, wa_nee_count, mapping

#%% Deleting all useless instances

# deleting speeders
FL = FL[FL['Duration (in seconds)'] >= 300]
WA = WA[WA['Duration (in seconds)'] >= 300]

# deleting people who don't keep cream in fridge
FL = FL[FL['Inclusion 4'] != 'Nooit']
FL = FL[FL['Inclusion 4'] != 'Non, jamais']
WA = WA[WA['Inclusion 4'] != 'Non, jamais']
WA = WA[WA['Inclusion 4'] != 'Nooit']

WA = WA.dropna(subset=['Inclusion 4'])
FL = FL.dropna(subset=['Inclusion 4'])


# deleting people who did not finish the survey
FL = FL[FL['Progress'] >= 100]
WA = WA[WA['Progress'] >= 100]

# deleting non Belgian residents1
FL = FL[FL['Q152'] != 'Nee']
FL = FL[FL['Q152'] != 'Non']
WA = WA[WA['Q152'] != 'Nee']
WA = WA[WA['Q152'] != 'Non']

WA = WA.dropna(subset=['Q152'])
FL = FL.dropna(subset=['Q152'])

# deleting people who don't use cooking cream
FL = FL[FL['Inclusion 3'] != 'Nooit']
FL = FL[FL['Inclusion 3'] != 'Non']
WA = WA[WA['Inclusion 3'] != 'Nooit']
WA = WA[WA['Inclusion 3'] != 'Non']

FL = FL.dropna(subset=['Inclusion 3'])
WA = WA.dropna(subset=['Inclusion 3'])

# deleting people who don't cook
FL = FL[FL['Q159'] != 'Nooit']
FL = FL[FL['Q159'] != 'Jamais']
WA = WA[WA['Q159'] != 'Jamais']
WA = WA[WA['Q159'] != 'Nooit']
WA = WA.dropna(subset=['Q159'])
FL = FL.dropna(subset=['Q159'])

# deleting people who don't consent
FL = FL[FL['Intro 1'] != 'Niet akkoord om deel te nemen']
FL = FL[FL['Intro 1'] != 'Non']
WA = WA[WA['Intro 1'] != 'Non']
WA = WA[WA['Intro 1'] != 'Niet akkoord om deel te nemen']
WA = WA.dropna(subset=['Intro 1'])
FL = FL.dropna(subset=['Intro 1'])

FL.reset_index(drop=True, inplace=True)
WA.reset_index(drop=True, inplace=True)

""" Deleting these columns """

columns_to_delete = ['Duration (in seconds)', 'Progress', 'Q152', 
                     'Inclusion 3', 'Q159','Q176_4_TEXT', 'ResponseId', 
                     'Intro 1']
FL = FL.drop(columns=columns_to_delete)
WA = WA.drop(columns=columns_to_delete)

del columns_to_delete


#%% Preprocessing of B1C1, ... , B3C6 columns

""" Ensuring uniform values. Option 1 is changed to 1, and option 2 is
changed to 2 """

replacement_mapping = {
    'Optie 1: Ik zou deze kookroom gebruiken.': 1,
    'Optie 2: Ik zou deze kookroom weggooien en een nieuw pakje nemen.': 2
}

columns_to_replace = ['B1C1', 'B1C2', 'B1C3', 'B1C4', 'B1C5', 'B1C6',
                      'B2C1', 'B2C2', 'B2C3', 'B2C4', 'B2C5', 'B2C6',
                      'B3C1', 'B3C2', 'B3C3', 'B3C4', 'B3C5', 'B3C6']

for column in columns_to_replace:
    FL[column] = FL[column].replace(replacement_mapping)
for column in columns_to_replace:
    WA[column] = WA[column].replace(replacement_mapping)
    
replacement_mapping = {
    'Option 1 : j\'utiliserais ce paquet de crème à cuisson liquide': 1,
    'Option 2 : je jetterais le paquet de crème à cuisson liquide et j\'en prendrais un nouveau': 2,
    'Option 2 : je jetterais le paquet de crème à cuisson liquide et j\'en prendrais un neuf': 2
}

columns_to_replace = ['B1C1', 'B1C2', 'B1C3', 'B1C4', 'B1C5', 'B1C6',
                      'B2C1', 'B2C2', 'B2C3', 'B2C4', 'B2C5', 'B2C6',
                      'B3C1', 'B3C2', 'B3C3', 'B3C4', 'B3C5', 'B3C6']

for column in columns_to_replace:
    WA[column] = WA[column].replace(replacement_mapping)
for column in columns_to_replace:
    FL[column] = FL[column].replace(replacement_mapping)
    
del replacement_mapping, columns_to_replace, column

#%% Add a column called Sets to FL and WA

""" Has the value 1 when the respondent was shown block 1 of choicecard"""

def calculate_sets(row):
    global deleted_rows
    sum_b1 = row[['B1C1', 'B1C2', 'B1C3', 'B1C4', 'B1C5', 'B1C6']].sum()
    sum_b2 = row[['B2C1', 'B2C2', 'B2C3', 'B2C4', 'B2C5', 'B2C6']].sum()
    sum_b3 = row[['B3C1', 'B3C2', 'B3C3', 'B3C4', 'B3C5', 'B3C6']].sum()

    if sum_b1 > 0:
        return 1
    elif sum_b2 > 0:
        return 2
    elif sum_b3 > 0:
        return 3
    else:
        return 4

FL['Sets'] = FL.apply(calculate_sets, axis=1)
WA['Sets'] = WA.apply(calculate_sets, axis=1)

""" People that have value 4 for this column provided no answers to the 
choice cards, so they unexplainably evaded the inclusion factors """

FL_sets_4_count = (FL['Sets'] == 4).sum()
WA_sets_4_count = (WA['Sets'] == 4).sum()

print("Rows with value 4 in 'Sets' column in FL:", FL_sets_4_count)
print("Rows with value 4 in 'Sets' column in WA:", WA_sets_4_count)

""" Deleting these people. """

FL = FL[FL['Sets'] != 4].reset_index(drop=True)
WA = WA[WA['Sets'] != 4].reset_index(drop=True)

del FL_sets_4_count, WA_sets_4_count

#%% Looking at irregularities in the Province column

""" How many Unknown provinces? """

FL_unknown_province_count = (FL['Province'] == 'Unknown').sum()
WA_unknown_province_count = (WA['Province'] == 'Unknown').sum()

print("Number of rows with 'Unknown' value for the column 'Province' in FL:", FL_unknown_province_count)
print("Number of rows with 'Unknown' value for the column 'Province' in WA:", WA_unknown_province_count)

del FL_unknown_province_count, WA_unknown_province_count

#%% Preprocessing 'AGE' 

""" Renaming the column """

FL.rename(columns={'Vraag 22': 'Age'}, inplace=True)
WA.rename(columns={'Vraag 22': 'Age'}, inplace=True)

""" Creating uniform values """

FL['Age'].replace({'55-64 jaar': '55-64', '45-54 jaar': '45-54', '35-44 jaar': '35-44', 
                   '25-34 jaar': '25-34', '18-24 jaar': '18-24'}, inplace=True)
WA['Age'].replace({'55-64 jaar': '55-64', '45-54 jaar': '45-54', '35-44 jaar': '35-44', 
                   '25-34 jaar': '25-34', '18-24 jaar': '18-24'}, inplace=True)

WA['Age'].replace({'de 55 à 64 ans': '55-64', '65 ans et plus': '65+', 'de 45 à 54 ans': '45-54', 
                   'de 35 à 44 ans': '35-44', 'de 25 à 34 ans': '25-34', 'de 18 à 24 ans': '18-24'}, inplace=True)
FL['Age'].replace({'de 55 à 64 ans': '55-64', '65 ans et plus': '65+', 'de 45 à 54 ans': '45-54', 
                   'de 35 à 44 ans': '35-44', 'de 25 à 34 ans': '25-34', 'de 18 à 24 ans': '18-24'}, inplace=True)

""" How mant unknown ages? """

FL_nan_age_count = FL['Age'].isna().sum()
WA_nan_age_count = WA['Age'].isna().sum()

print("Number of rows with NaN values for the column 'Age' in FL:", FL_nan_age_count)
print("Number of rows with NaN values for the column 'Age' in WA:", WA_nan_age_count)

del WA_nan_age_count, FL_nan_age_count

#%%  Preprocessing 'GENDER'

""" Changing the name of the column """

FL.rename(columns={'Q176': 'Gender'}, inplace=True)
WA.rename(columns={'Q176': 'Gender'}, inplace=True)

""" Creating uniform values"""

gender_mapping = {'Homme': 1, 'Femme': 0, 'Man': 1, 'Vrouw': 0}

# Replace values in the 'Gender' column of FL and WA
FL['Gender'] = FL['Gender'].replace(gender_mapping)
WA['Gender'] = WA['Gender'].replace(gender_mapping)

# Replace any other values with NaN
FL['Gender'] = FL['Gender'].where(FL['Gender'].isin([0, 1]), np.nan)
WA['Gender'] = WA['Gender'].where(WA['Gender'].isin([0, 1]), np.nan)

""" How many unkown genders? """

FL_nan_age_count = FL['Gender'].isna().sum()
WA_nan_age_count = WA['Gender'].isna().sum()

print("Number of rows with NaN values for the column 'Gender' in FL:", FL_nan_age_count)
print("Number of rows with NaN values for the column 'Gender' in WA:", WA_nan_age_count)

del WA_nan_age_count, FL_nan_age_count, gender_mapping

#%% Preprocessing 'INCOME'

""" changing the name of the column """

FL.rename(columns={'Vraag 27 ': 'Income per month'}, inplace=True)
WA.rename(columns={'Vraag 27 ': 'Income per month'}, inplace=True)

""" Making both values sets uniform """

FL['Income per month'] = FL['Income per month'].str.replace(' per maand', '')
FL['Income per month'] = FL['Income per month'].str.replace(' par mois', '')
FL['Income per month'] = FL['Income per month'].str.replace('par moi', '')

WA['Income per month'] = WA['Income per month'].str.replace(' par mois', '')
WA['Income per month'] = WA['Income per month'].str.replace('par moi', '')
WA['Income per month'] = WA['Income per month'].str.replace(' per maand', '')

#%% Preprocessing 'Household tally'

""" changing the name of the column """

FL.rename(columns={'Vraag 24a': 'Household tally'}, inplace=True)
WA.rename(columns={'Vraag 24a': 'Household tally'}, inplace=True)

""" changing extreme impossible values with 0 """

FL['Household tally'] = pd.to_numeric(FL['Household tally'], errors='coerce')
FL['Household tally'] = np.where(FL['Household tally'] > 7, np.nan, FL['Household tally'])
WA['Household tally'] = pd.to_numeric(WA['Household tally'], errors='coerce')
WA['Household tally'] = np.where(WA['Household tally'] > 7, np.nan, WA['Household tally'])

""" Changing 0 values to 1 """

FL_zero_to_one_count = (FL['Household tally'] == 0).sum()
WA_zero_to_one_count = (WA['Household tally'] == 0).sum()

FL['Household tally'].replace(0, 1, inplace=True)
WA['Household tally'].replace(0, 1, inplace=True)

print("Changed 0 values to 1 in FL:", FL_zero_to_one_count)
print("Changed 0 values to 1 in WA:", WA_zero_to_one_count)

""" How many unknown household tallies """

FL_nan_household_count = FL['Household tally'].isna().sum()
WA_nan_household_count = WA['Household tally'].isna().sum()

print("Number of instances with NaN values for the column 'Household tally' in FL:", FL_nan_household_count)
print("Number of instances with NaN values for the column 'Household tally' in WA:", WA_nan_household_count)

del FL_nan_household_count, FL_zero_to_one_count, WA_nan_household_count, WA_zero_to_one_count

#%% Preprocessing 'Years in Belgium'

""" changing the name of the column """

FL.rename(columns={'Vraag 25': 'Years in Belgium'}, inplace=True)
WA.rename(columns={'Vraag 25': 'Years in Belgium'}, inplace=True)

""" Ensuring uniform values """

mapping_dict = {
    'Ik heb altijd in België gewoond': 'Always',
    'Meer dan 10 jaar': '> 10',
    '5 jaar of minder': '5 or less',
    '10 jaar of minder': '10 or less',
    '1 jaar of minder': '1 or less',
    'Plus de 10 ans': '> 10',
    "J'ai toujours vécu en Belgique": 'Always',
    '5 ans ou moins': '5 or less',
    '10 ans ou moins': '10 or less',
    '1 an ou moins': '1 or less'
}
FL['Years in Belgium'] = FL['Years in Belgium'].replace(mapping_dict)
WA['Years in Belgium'] = WA['Years in Belgium'].replace(mapping_dict)

del mapping_dict

#%% Preprocessing education

""" changing the name of the column """

FL.rename(columns={'Vraag 26': 'Education'}, inplace=True)
WA.rename(columns={'Vraag 26': 'Education'}, inplace=True)

""" Uniform values """

mapping_dict_education = {
    'Secundair onderwijs': 'Secondary Education',
    'Primair onderwijs': 'Primary Education',
    'Geen': 'None',
    'PhD, Doctoraat': 'PhD'  
}
FL['Education'] = FL['Education'].replace(mapping_dict_education)
WA['Education'] = WA['Education'].replace(mapping_dict_education)


mapping_dict_wa_education = {
    'Secondaire': 'Secondary Education',
    'Licence': 'Bachelor',
    'Ph.D., Doctorat': 'PhD',
    'Primaire': 'Primary Education',
    'Ne se prononce pas': 'None',
    'Aucun': 'None'
}
WA['Education'] = WA['Education'].replace(mapping_dict_wa_education)
FL['Education'] = FL['Education'].replace(mapping_dict_wa_education)

""" How many unknown educations? """

FL_none_education_count = (FL['Education'] == 'None').sum()
WA_none_education_count = (WA['Education'] == 'None').sum()

print("Number of instances with the value 'None' for the column 'Education' in FL:", FL_none_education_count)
print("Number of instances with the value 'None' for the column 'Education' in WA:", WA_none_education_count)

del FL_none_education_count, mapping_dict_education, mapping_dict_wa_education, WA_none_education_count

#%% Now deleting all the unknown demographics

# unknown province
FL = FL[FL['Province'] != 'Unknown'].reset_index(drop=True)
WA = WA[WA['Province'] != 'Unknown'].reset_index(drop=True)

# unknown gender
FL = FL.dropna(subset=['Gender']).reset_index(drop=True)
WA = WA.dropna(subset=['Gender']).reset_index(drop=True)

# unkown age
FL = FL.dropna(subset=['Age']).reset_index(drop=True)
WA = WA.dropna(subset=['Age']).reset_index(drop=True)

# impossible or unkown household tally
FL = FL.dropna(subset=['Household tally']).reset_index(drop=True)
WA = WA.dropna(subset=['Household tally']).reset_index(drop=True)

# unkown or impossible education
FL = FL[FL['Education'] != 'None'].reset_index(drop=True)
WA = WA[WA['Education'] != 'None'].reset_index(drop=True)

#%% Checking quotas

""" Checking what the current representation is of every demographic group
outlined in the quotas which were given to Dynata as a target """

FL_counts = {
    'Gender:0 & Age:18-24': len(FL[(FL['Gender'] == 0) & (FL['Age'] == '18-24')]),
    'Gender:0 & Age:25-34': len(FL[(FL['Gender'] == 0) & (FL['Age'] == '25-34')]),
    'Gender:0 & Age:35-44': len(FL[(FL['Gender'] == 0) & (FL['Age'] == '35-44')]),
    'Gender:0 & Age:45-54': len(FL[(FL['Gender'] == 0) & (FL['Age'] == '45-54')]),
    'Gender:0 & Age:55-64': len(FL[(FL['Gender'] == 0) & (FL['Age'] == '55-64')]),
    'Gender:0 & Age:65+': len(FL[(FL['Gender'] == 0) & (FL['Age'] == '65+')]),
    'Gender:1 & Age:18-24': len(FL[(FL['Gender'] == 1) & (FL['Age'] == '18-24')]),
    'Gender:1 & Age:25-34': len(FL[(FL['Gender'] == 1) & (FL['Age'] == '25-34')]),
    'Gender:1 & Age:35-44': len(FL[(FL['Gender'] == 1) & (FL['Age'] == '35-44')]),
    'Gender:1 & Age:45-54': len(FL[(FL['Gender'] == 1) & (FL['Age'] == '45-54')]),
    'Gender:1 & Age:55-64': len(FL[(FL['Gender'] == 1) & (FL['Age'] == '55-64')]),
    'Gender:1 & Age:65+': len(FL[(FL['Gender'] == 1) & (FL['Age'] == '65+')]),
    'Income: €1001 - €2000': len(FL[FL['Income per month'] == '€1001 – €2000']),
    'Income: €2001 - €3000': len(FL[FL['Income per month'] == '€2001 – €3000']),
    'Income: €3001 - €4000': len(FL[FL['Income per month'] == '€3001 – €4000']),
    'Income: €4001 - €5000': len(FL[FL['Income per month'] == '€4001 – €5000']),
    'Income: > €5000': len(FL[FL['Income per month'] == '> €5000'])
}

WA_counts = {
    'Gender:0 & Age:18-24': len(WA[(WA['Gender'] == 0) & (WA['Age'] == '18-24')]),
    'Gender:0 & Age:25-34': len(WA[(WA['Gender'] == 0) & (WA['Age'] == '25-34')]),
    'Gender:0 & Age:35-44': len(WA[(WA['Gender'] == 0) & (WA['Age'] == '35-44')]),
    'Gender:0 & Age:45-54': len(WA[(WA['Gender'] == 0) & (WA['Age'] == '45-54')]),
    'Gender:0 & Age:55-64': len(WA[(WA['Gender'] == 0) & (WA['Age'] == '55-64')]),
    'Gender:0 & Age:65+': len(WA[(WA['Gender'] == 0) & (WA['Age'] == '65+')]),
    'Gender:1 & Age:18-24': len(WA[(WA['Gender'] == 1) & (WA['Age'] == '18-24')]),
    'Gender:1 & Age:25-34': len(WA[(WA['Gender'] == 1) & (WA['Age'] == '25-34')]),
    'Gender:1 & Age:35-44': len(WA[(WA['Gender'] == 1) & (WA['Age'] == '35-44')]),
    'Gender:1 & Age:45-54': len(WA[(WA['Gender'] == 1) & (WA['Age'] == '45-54')]),
    'Gender:1 & Age:55-64': len(WA[(WA['Gender'] == 1) & (WA['Age'] == '55-64')]),
    'Gender:1 & Age:65+': len(WA[(WA['Gender'] == 1) & (WA['Age'] == '65+')]),
    'Income: €1001 - €2000': len(WA[WA['Income per month'] == '€1001 – €2000']),
    'Income: €2001 - €3000': len(WA[WA['Income per month'] == '€2001 – €3000']),
    'Income: €3001 - €4000': len(WA[WA['Income per month'] == '€3001 – €4000']),
    'Income: €4001 - €5000': len(WA[WA['Income per month'] == '€4001 – €5000']),
    'Income: > €5000': len(WA[WA['Income per month'] == '> €5000'])
}

df_counts = pd.DataFrame({'FL': FL_counts, 'WA': WA_counts})
print(df_counts)
df_counts.to_excel('Quotas.xlsx', index=True)

#%% Deleting people to meet quota's

""" See seperate script on how this was done """

""" To avoid randomly picking people to exclude because of the quota targets,
one randomization was done using the seperate script and then the IDs were 
saved """

Random_deletions_FL = ['1,336', '1,279', '1,127', '2,905', '2,1096', '1,312', 
                       '1,1058', '1,530', '1,767', '1,627', '1,533', '1,1086', 
                       '1,1002', '1,476', '1,666', '1,394', '1,637', '1,362', 
                       '1,1262', '2,896', '1,490', '1,1122', '2,999', '2,1043', 
                       '1,752', '2,1075', '1,586', '1,497', '1,651', '1,670', 
                       '1,326', '1,369', '1,34', '1,492', '1,116', '1,474', 
                       '2,1301', '2,1073', '1,548', '1,415', '1,576', '1,1317', 
                       '2,1130', '1,578', '1,945', '1,230', '1,320', '1,151', 
                       '1,145', '2,1014', '1,660', '1,1269', '1,1241', '1,765', 
                       '2,1080', '1,690', '1,247', '1,612', '1,680', '1,581', 
                       '1,513', '2,34', '1,417', '1,798', '1,1281', '2,937', 
                       '1,847', '1,1386', '1,613', '1,545', '1,1092', '2,1082', 
                       '1,355', '1,112', '1,475', '2,1342', '1,1302', '1,866', 
                       '1,1416', '1,305', '1,1400', '2,1035', '2,1244', 
                       '2,1288', '1,718', '1,357', '1,1339', '1,1381', '1,126', 
                       '1,1245', '1,506', '1,650', '1,149', '1,720', '1,315', 
                       '1,179', '1,537', '1,549', '1,579', '1,67', '1,223', 
                       '1,1412', '2,1139', '1,109', '1,1380', '1,704'
                       , '2,1239', '1,943', '2,1085', '2,907', '1,595', 
                       '1,1496', '1,1485', '1,1493', '2,1263', '1,938', 
                       '1,671', '1,388', '1,306', '1,185', '1,800', '1,253', 
                       '1,143', '1,113', '1,96', '1,224', '1,1399', '1,298', 
                       '1,1034', '1,52', '1,777', '1,360', '1,1303', '2,1160', 
                       '1,62', '1,753', '1,532', '1,1052', '1,1312', '1,830', 
                       '1,713', '1,1054', '1,400', '1,435', '1,1270', '1,1351', 
                       '1,465', '2,1016', '1,1323', '1,1308', '2,1220', 
                       '2,938', '2,959', '1,1325', '1,153', '2,1318', '1,265', 
                       '1,839', '1,1163', '1,1284', '2,939', '1,610', '1,747', 
                       '1,1417', '1,77', '1,344', '1,1419', '1,222', '2,1148', 
                       '1,534', '2,40', '1,683', '1,266', '1,124', '1,268', 
                       '1,1249', '1,793', '1,624', '1,49', '1,772', '1,567', 
                       '1,662', '1,319', '1,44', '1,439', '1,456', '1,494', 
                       '1,649', '1,56', '1,600', '1,13', '1,781', '2,1058', 
                       '1,286', '1,515', '1,169', '1,297', '1,674', '2,1101', 
                       '1,174', '1,42', '2,1039', '2,1156', '2,1145', '1,74', 
                       '2,1113', '1,509', '1,551', '1,748', '2,1116', '1,689', 
                       '2,1044', '1,1263', '1,437', '1,32', '1,21', '1,1426', 
                       '1,951', '1,239', '2,1336', '1,569', '1,850', '1,134', 
                       '1,536', '1,734', '1,1393', '1,799', '1,529', '1,128', 
                       '2,1258', '1,561', '1,849', '1,45', '1,1346', '2,1261', 
                       '1,1388', '1,844', '2,1354', '1,914', '1,698', '1,397', 
                       '1,1335', '1,1415', '1,1299', '1,500', '2,1278', 
                       '1,1368', '1,1347', '1,601', '2,978', '1,1384', 
                       '1,1354', '1,1418', '1,1370', '2,1381', '2,969', 
                       '1,661', '1,55']
Random_deletions_WA = ['2,727', '2,54', '2,441', '2,90', '2,172', '2,1367', 
                       '2,327', '2,88', '2,1203', '2,128', '2,352', '2,540', 
                       '2,391', '2,125', '2,37', '2,297', '2,515', '2,268', 
                       '2,453', '2,449', '2,1353', '1,71', '2,512', '2,403', 
                       '2,334', '2,337', '2,1275', '2,374', '2,91', '2,370', 
                       '2,39', '2,131', '2,832', '1,924', '2,130', '2,164', 
                       '2,33', '2,86', '2,701', '2,133', '2,5', '2,346', 
                       '2,402']

FL = FL[~FL['ID'].isin(Random_deletions_FL)].reset_index(drop=True)
WA = WA[~WA['ID'].isin(Random_deletions_WA)].reset_index(drop=True)

del FL_counts, WA_counts, Random_deletions_FL, Random_deletions_WA

#%% New quotas

FL_counts_new = {
    'Gender:0 & Age:18-24': len(FL[(FL['Gender'] == 0) & (FL['Age'] == '18-24')]),
    'Gender:0 & Age:25-34': len(FL[(FL['Gender'] == 0) & (FL['Age'] == '25-34')]),
    'Gender:0 & Age:35-44': len(FL[(FL['Gender'] == 0) & (FL['Age'] == '35-44')]),
    'Gender:0 & Age:45-54': len(FL[(FL['Gender'] == 0) & (FL['Age'] == '45-54')]),
    'Gender:0 & Age:55-64': len(FL[(FL['Gender'] == 0) & (FL['Age'] == '55-64')]),
    'Gender:0 & Age:65+': len(FL[(FL['Gender'] == 0) & (FL['Age'] == '65+')]),
    'Gender:1 & Age:18-24': len(FL[(FL['Gender'] == 1) & (FL['Age'] == '18-24')]),
    'Gender:1 & Age:25-34': len(FL[(FL['Gender'] == 1) & (FL['Age'] == '25-34')]),
    'Gender:1 & Age:35-44': len(FL[(FL['Gender'] == 1) & (FL['Age'] == '35-44')]),
    'Gender:1 & Age:45-54': len(FL[(FL['Gender'] == 1) & (FL['Age'] == '45-54')]),
    'Gender:1 & Age:55-64': len(FL[(FL['Gender'] == 1) & (FL['Age'] == '55-64')]),
    'Gender:1 & Age:65+': len(FL[(FL['Gender'] == 1) & (FL['Age'] == '65+')]),
    'Income: €1001 - €2000': len(FL[FL['Income per month'] == '€1001 – €2000']),
    'Income: €2001 - €3000': len(FL[FL['Income per month'] == '€2001 – €3000']),
    'Income: €3001 - €4000': len(FL[FL['Income per month'] == '€3001 – €4000']),
    'Income: €4001 - €5000': len(FL[FL['Income per month'] == '€4001 – €5000']),
    'Income: > €5000': len(FL[FL['Income per month'] == '> €5000'])
}

WA_counts_new = {
    'Gender:0 & Age:18-24': len(WA[(WA['Gender'] == 0) & (WA['Age'] == '18-24')]),
    'Gender:0 & Age:25-34': len(WA[(WA['Gender'] == 0) & (WA['Age'] == '25-34')]),
    'Gender:0 & Age:35-44': len(WA[(WA['Gender'] == 0) & (WA['Age'] == '35-44')]),
    'Gender:0 & Age:45-54': len(WA[(WA['Gender'] == 0) & (WA['Age'] == '45-54')]),
    'Gender:0 & Age:55-64': len(WA[(WA['Gender'] == 0) & (WA['Age'] == '55-64')]),
    'Gender:0 & Age:65+': len(WA[(WA['Gender'] == 0) & (WA['Age'] == '65+')]),
    'Gender:1 & Age:18-24': len(WA[(WA['Gender'] == 1) & (WA['Age'] == '18-24')]),
    'Gender:1 & Age:25-34': len(WA[(WA['Gender'] == 1) & (WA['Age'] == '25-34')]),
    'Gender:1 & Age:35-44': len(WA[(WA['Gender'] == 1) & (WA['Age'] == '35-44')]),
    'Gender:1 & Age:45-54': len(WA[(WA['Gender'] == 1) & (WA['Age'] == '45-54')]),
    'Gender:1 & Age:55-64': len(WA[(WA['Gender'] == 1) & (WA['Age'] == '55-64')]),
    'Gender:1 & Age:65+': len(WA[(WA['Gender'] == 1) & (WA['Age'] == '65+')]),
    'Income: €1001 - €2000': len(WA[WA['Income per month'] == '€1001 – €2000']),
    'Income: €2001 - €3000': len(WA[WA['Income per month'] == '€2001 – €3000']),
    'Income: €3001 - €4000': len(WA[WA['Income per month'] == '€3001 – €4000']),
    'Income: €4001 - €5000': len(WA[WA['Income per month'] == '€4001 – €5000']),
    'Income: > €5000': len(WA[WA['Income per month'] == '> €5000'])
}

df_counts_new = pd.DataFrame({'FL': FL_counts_new, 'WA': WA_counts_new})
print(df_counts_new)
df_counts_new.to_excel('Quotas_new.xlsx', index=True)

""" number of deletions made to meet target representation """

deleted = df_counts.iloc[:12] - df_counts_new.iloc[:12]
print(deleted.values.sum())

del deleted, df_counts, df_counts_new, FL_counts_new, WA_counts_new

#%% Returning the psids of the final dataset

""" These are the IDs of all the people who will be used for analysis """

PSIDs = []

for id_value in FL['ID']:
    if id_value.startswith('1'):
        response_id = OG_FL.loc[OG_FL['ID'] == id_value, 'psid'].values
        if len(response_id) > 0:
            PSIDs.extend(response_id)

    elif id_value.startswith('2'):
        response_id = OG_WA.loc[OG_WA['ID'] == id_value, 'psid'].values
        if len(response_id) > 0:
            PSIDs.extend(response_id)
            
for id_value in WA['ID']:
    if id_value.startswith('1'):
        response_id = OG_FL.loc[OG_FL['ID'] == id_value, 'psid'].values
        if len(response_id) > 0:
            PSIDs.extend(response_id)

    elif id_value.startswith('2'):
        response_id = OG_WA.loc[OG_WA['ID'] == id_value, 'psid'].values
        if len(response_id) > 0:
            PSIDs.extend(response_id)
            
PSIDs_df = pd.DataFrame({'psid': PSIDs})
PSIDs_df.to_excel('PSIDs.xlsx', index=False)

del id_value, PSIDs, PSIDs_df, response_id

#%% 

""" NOW ALL RESPONDENTS ARE CLEARED FOR ANALYSIS """

""" Protest responses, people who indicated the opt out each time and 
people who indicated that they didn't find the choicecards clear should now be deleted.
Yet, these responses are results in and of themselves, so they should not be replaced
by respondents which were left out during the randomized deletion phase to meet
the target quotas """

#%% Demographics stats and figures

total = pd.concat([FL, WA], ignore_index=True)

""" Gender """

total_instances = len(total)
freq_1 = total['Gender'].value_counts(normalize=True).get(1, 0) * 100  
freq_0 = total['Gender'].value_counts(normalize=True).get(0, 0) * 100 

print(f"Relative frequency of males: {freq_1:.2f}%")
print(f"Relative frequency of females: {freq_0:.2f}%")

""" Province """

total_instances = len(total)
province_freq = total['Province'].value_counts(normalize=True) * 100  

print("Relative frequencies of unique values in the 'Province' column:")
for province, freq in province_freq.items():
    print(f"{province}: {freq:.2f}%")

""" Age """

age_freq = total['Age'].value_counts(normalize=True) * 100  

print("Relative frequencies of unique values in the 'Age' column:")
for age, freq in age_freq.items():
    print(f"Age {age}: {freq:.2f}%")

""" Income per month """

income_freq = total['Income per month'].value_counts(normalize=True) * 100  

print("Relative frequencies of unique values in the 'Income per month' column:")
for income, freq in income_freq.items():
    print(f"Income per month {income}: {freq:.2f}%")
    
""" Household tally """
household_freq = total['Household tally'].value_counts(normalize=True) * 100  

print("Relative frequencies of unique values in the 'Household tally' column:")
for household, freq in household_freq.items():
    print(f"Household tally {household}: {freq:.2f}%")
    
average_household_tally = total['Household tally'].mean()

print(f"The average of the 'Household tally' column is: {average_household_tally:.2f}")

""" Years in Belgium """

years_in_belgium_freq = total['Years in Belgium'].value_counts(normalize=True) * 100   
print("Relative frequencies of unique values in the 'Years in Belgium' column:")
for years, freq in years_in_belgium_freq.items():
    print(f"Years in Belgium {years}: {freq:.2f}%")
    
""" Education """

education_freq = total['Education'].value_counts(normalize=True) * 100  

print("Relative frequencies of unique values in the 'Education' column:")
for education, freq in education_freq.items():
    print(f"Education '{education}': {freq:.2f}%")
    
del age, age_freq, average_household_tally, education, education_freq, freq, freq_0, freq_1
del household, household_freq, income, income_freq, total_instances, total, years, years_in_belgium_freq
del province, province_freq

#%% Understanding the column Vraag 16a  - Whether the choice cards were clear

""" Uniform values """

def check_condition(value):
    if pd.isna(value):
        return 0
    if 'Nee, niet altijd' in value or "No" in value:
        return 0
    return 1

FL.insert(49, 'Clarity', FL['Vraag 16a'].apply(lambda x: check_condition(x)))
WA.insert(49, 'Clarity', WA['Vraag 16a'].apply(lambda x: check_condition(x)))

""" Checking frequencies """

combined_df = pd.concat([FL, WA], ignore_index=True)
print(combined_df['Clarity'].value_counts(normalize=True) * 100)  

""" Deleting the people for whom the choice cards were unclear """

FL = FL[FL['Clarity'] != 0]
WA = WA[WA['Clarity'] != 0]
FL = FL.reset_index(drop=True)
WA = WA.reset_index(drop=True)

""" Dropping the column """

FL.drop(columns=['Vraag 16a'], inplace=True)
WA.drop(columns=['Vraag 16a'], inplace=True)
FL.drop(columns=['Vraag 16b '], inplace=True)
WA.drop(columns=['Vraag 16b '], inplace=True)
FL.drop(columns=['Clarity'], inplace=True)
WA.drop(columns=['Clarity'], inplace=True)

#%% Understanding column Vraag 17 and Vraag 17 _7_TEXT: why they chose the opt out each time

# FL: 6
#	Er waren te veel factoren om te overwegen
#	Ik heb meer informatie nodig om een mening te vormen
#	Ik heb meer informatie nodig om een mening te vormen
#	J'ai besoin de plus d'informations pour me faire une opinion
# Il y avait trop de facteurs à prendre en compte
# Ik hou me niet bezig met voedselverspilling dus het maakt voor mij niet uit

# WA: 6
#	J'ai besoin de plus d'informations pour me faire une opinion
#	Je ne me préoccupe pas du gaspillage alimentaire, cela ne fait donc aucune différence pour moi
#	Je ne me préoccupe pas du gaspillage alimentaire, cela ne fait donc aucune différence pour moi
# Je pense qu'un produit ouvert doit être consommé dans les 2-3 jours
# Ik hou me niet bezig met voedselverspilling dus het maakt voor mij niet uit
# C'était écrit en néerlandais : je me suis peut-être trompé

""" 2 protestors """

# deleting people who chose opt out each time

FL = FL[pd.isnull(FL['Vraag 17 '])]
WA = WA[pd.isnull(WA['Vraag 17 '])]
FL = FL.reset_index(drop=True)
WA = WA.reset_index(drop=True)

# Deleting columns

FL.drop(columns=['Vraag 17 '], inplace=True)
WA.drop(columns=['Vraag 17 '], inplace=True)

FL.drop(columns=['Vraag 17 _7_TEXT'], inplace=True)
WA.drop(columns=['Vraag 17 _7_TEXT'], inplace=True)

#%% MIDWAYPOINT

""" NOW ALL THE DELETIONS HAVE BEEN MADE """
""" These deletions however should not be made before reporting on the proportions
of the answers of the other questions in the survey """

#%% Create the DCE dataset wih every unique ID repeated 12 times (2 choices per 6 cards)

DCE_FL = pd.DataFrame()
repeated_ids = np.repeat(FL['ID'], 12)
DCE_FL['ID'] = repeated_ids
DCE_FL['Region'] = 1 # 1 is Flanders

DCE_WA = pd.DataFrame()
repeated_ids = np.repeat(WA['ID'], 12)
DCE_WA['ID'] = repeated_ids
DCE_WA['Region'] = 0 # 0 is Wallonia

del repeated_ids 


#%% Alter
""" This column has values 1 and 2 relecting the possible choices"""

pattern_FL = [(i % 2) + 1 for i in range(len(DCE_FL))]
pattern_WA = [(i % 2) + 1 for i in range(len(DCE_WA))]
DCE_FL['Alter'] = pattern_FL
DCE_WA['Alter'] = pattern_WA

del pattern_FL, pattern_WA


#%% Alter3 
""" This column has the value for the opt out option which is throwing away
the cooking cream """

sequence = [0, 1, 0, 1]
repeated_sequence_FL = sequence * (len(DCE_FL) // len(sequence)) + sequence[:len(DCE_FL) % len(sequence)]
repeated_sequence_WA = sequence * (len(DCE_WA) // len(sequence)) + sequence[:len(DCE_WA) % len(sequence)]
DCE_FL['Alter3'] = repeated_sequence_FL
DCE_WA['Alter3'] = repeated_sequence_WA
del sequence, repeated_sequence_FL, repeated_sequence_WA

#%% Chtaskid
""" This column reflects the number of choice cards"""

sequence_FL = [i for i in range(1, len(DCE_FL) // 2 + 2) for _ in range(2)]
sequence_WA = [i for i in range(1, len(DCE_WA) // 2 + 2) for _ in range(2)]

sequence_FL = sequence_FL[:len(DCE_FL)]
sequence_WA = sequence_WA[:len(DCE_WA)]

DCE_FL['Chtaskid'] = sequence_FL
DCE_WA['Chtaskid'] = [seq + len(DCE_FL)/2 for seq in sequence_WA]

del sequence_FL, sequence_WA

#%% Choicesets
""" Reflects the choicecard in the set. There are 6 cards per set. """

sequence = [i for i in range(1, 7) for _ in range(2)]
repeated_sequence_FL = sequence * (len(DCE_FL) // len(sequence)) + sequence[:len(DCE_FL) % len(sequence)]
repeated_sequence_WA = sequence * (len(DCE_WA) // len(sequence)) + sequence[:len(DCE_WA) % len(sequence)]

DCE_FL['Choicesets'] = repeated_sequence_FL
DCE_WA['Choicesets'] = repeated_sequence_WA

del sequence, repeated_sequence_WA, repeated_sequence_FL

#%% Set 
""" This column reflects the set of cards that the instance was shown"""

DCE_FL['Sets'] = 0

for index, row in DCE_FL.iterrows():
    id_value = row['ID']
    sets_value = FL.loc[FL['ID'] == id_value, 'Sets'].values[0]
    DCE_FL.at[index, 'Sets'] = sets_value
    
DCE_WA['Sets'] = 0

for index, row in DCE_WA.iterrows():
    id_value = row['ID']
    sets_value = WA.loc[WA['ID'] == id_value, 'Sets'].values[0]
    DCE_WA.at[index, 'Sets'] = sets_value
    
del id_value, index, row, sets_value
    
#%% Chosenalter

def calculate_chosen_alter(row): 
    set_col = f"B{row['Sets']}C{row['Choicesets']}" 
    return FL.loc[FL['ID'] == row['ID'], set_col].iloc[0] # Add Chosenalter column to 
DCE_FL['Chosenalter'] = DCE_FL.apply(calculate_chosen_alter, axis=1)

def calculate_chosen_alter(row): 
    set_col = f"B{row['Sets']}C{row['Choicesets']}" 
    return WA.loc[WA['ID'] == row['ID'], set_col].iloc[0] # Add Chosenalter column to 
DCE_WA['Chosenalter'] = DCE_WA.apply(calculate_chosen_alter, axis=1)

#%% Integrating the Flemish and the Walloon DCE dataset

DCE = pd.concat([DCE_FL, DCE_WA], ignore_index=True)

del DCE_FL, DCE_WA

#%% Coding the attributes

columns = ['NOV', 'DEC', 'JUN', 'LABEL_1', 'LABEL_2', 'OPEN_0','OPEN_2','OPEN_3','OPEN_7', 'PLASTIC', 'CAN', 'CARTON']
attributes = pd.DataFrame(columns=columns)

DCE[columns] = 0 

# Set 1

cards_1 = attributes.copy()
cards_1.loc[len(cards_1)] = [1,0,0,1,0,1,0,0,0,0,1,0]
cards_1.loc[len(cards_1)] = [0,1,0,1,0,1,0,0,0,0,1,0]
cards_1.loc[len(cards_1)] = [1,0,0,0,1,0,0,0,1,0,1,0]
cards_1.loc[len(cards_1)] = [0,1,0,1,0,0,0,0,1,1,0,0]
cards_1.loc[len(cards_1)] = [0,0,1,0,1,0,1,0,0,0,1,0]
cards_1.loc[len(cards_1)] = [0,0,1,0,1,0,0,0,1,0,0,1]

# Set 2

cards_2 = attributes.copy()
cards_2.loc[len(cards_2)] = [0,1,0,0,1,0,0,0,1,0,1,0]
cards_2.loc[len(cards_2)] = [1,0,0,0,1,1,0,0,0,1,0,0]
cards_2.loc[len(cards_2)] = [0,1,0,1,0,0,0,1,0,0,0,1]
cards_2.loc[len(cards_2)] = [1,0,0,1,0,0,0,1,0,1,0,0]
cards_2.loc[len(cards_2)] = [0,0,1,0,1,0,0,1,0,1,0,0]
cards_2.loc[len(cards_2)] = [1,0,0,0,1,0,1,0,0,0,0,1]

# Set 3

cards_3 = attributes.copy()
cards_3.loc[len(cards_3)] = [1,0,0,1,0,0,0,1,0,0,0,1]
cards_3.loc[len(cards_3)] = [0,0,1,1,0,0,1,0,0,0,0,1]
cards_3.loc[len(cards_3)] = [0,1,0,0,1,1,0,0,0,1,0,0]
cards_3.loc[len(cards_3)] = [0,0,1,1,0,0,0,0,1,1,0,0]
cards_3.loc[len(cards_3)] = [0,1,0,0,1,0,1,0,0,0,0,1]
cards_3.loc[len(cards_3)] = [0,0,1,1,0,0,0,1,0,0,1,0]

#%% Adding the attributes

for index, row in DCE.iterrows():
    set_value = row['Sets']
    option = row['Alter3']
    choice_set = row['Choicesets']
    if set_value == 1 and option != 1:
        DCE.loc[index, ['NOV', 'DEC', 'JUN', 'LABEL_1', 'LABEL_2', 'OPEN_0','OPEN_2',
                        'OPEN_3','OPEN_7', 'PLASTIC', 'CAN', 'CARTON']] = cards_1.iloc[choice_set - 1].values
    elif set_value == 2 and option != 1:
        DCE.loc[index, ['NOV', 'DEC', 'JUN', 'LABEL_1', 'LABEL_2', 'OPEN_0','OPEN_2',
                        'OPEN_3','OPEN_7', 'PLASTIC', 'CAN', 'CARTON']] = cards_2.iloc[choice_set - 1].values
    elif set_value == 3 and option != 1:
        DCE.loc[index, ['NOV', 'DEC', 'JUN', 'LABEL_1', 'LABEL_2', 'OPEN_0','OPEN_2','OPEN_3','OPEN_7', 'PLASTIC', 'CAN', 'CARTON']] = cards_3.iloc[choice_set - 1].values
        
del attributes, cards_1, cards_2, cards_3, choice_set, columns
del index, row, set_value, option

#%% Adding the variable Choice

""" Returns 1 when chosenalter is equel to alter and zero if not """

DCE.insert(4, 'Choice', 0)
DCE.loc[DCE['Alter'] == DCE['Chosenalter'], 'Choice'] = 1

#%%

""" Conjoining the column Q138 and Q138_7_TEXT """

FL.loc[FL['Q138'] == 'Andere, namelijk ...', 'Q138'] = FL.loc[FL['Q138'] == 'Andere, namelijk ...', 'Q138_7_TEXT']
FL.loc[FL['Q138'] == 'Autre, précisez...', 'Q138'] = FL.loc[FL['Q138'] == 'Autre, précisez...', 'Q138_7_TEXT']

WA.loc[WA['Q138'] == 'Andere, namelijk ...', 'Q138'] = WA.loc[WA['Q138'] == 'Andere, namelijk ...', 'Q138_7_TEXT']
WA.loc[WA['Q138'] == 'Autre, précisez...', 'Q138'] = WA.loc[WA['Q138'] == 'Autre, précisez...', 'Q138_7_TEXT']


""" Renaming the column """

FL.rename(columns={'Q138': 'Diet'}, inplace=True)
WA.rename(columns={'Q138': 'Diet'}, inplace=True)


""" Uniform values """

value_mapping = {
    'Ik heb geen dieetbeperkingen': 'None',
    'Vegetarisch (geen vlees en vis)': 'Vegetarian',
    'Vegan (geen dierlijke producten, inclusief zuivel en eieren)': 'Vegan',
    'Lactose intolerant (ik kan geen producten met lactose consumeren)': 'Lactose intolerant',
    'Pescotariër (vegetariër, maar ik eet vis/zeevruchten)': 'Pescatarian',
    'Flexitariër (vooral vegetarisch maar soms eet ik wel vlees/vis)': 'Flexitarian',
    'Ik heb geen dieetbeperkingen,Flexitariër (vooral vegetarisch maar soms eet ik wel vlees/vis)': 'Flexitarian',
    'Ik heb geen dieetbeperkingen,Pescotariër (vegetariër, maar ik eet vis/zeevruchten),Flexitariër (vooral vegetarisch maar soms eet ik wel vlees/vis)': 'Flexitarian & Pescatarian',
    'Ik heb geen dieetbeperkingen,Andere, namelijk ...': 'None',
    'Vegan (geen dierlijke producten, inclusief zuivel en eieren),Flexitariër (vooral vegetarisch maar soms eet ik wel vlees/vis)': 'Flexitarian',
    'Allergisch aan vis maar niet aan schaaldieren': 'No fish',
    'suikerarm': 'Low sugar',
    'Suikerarm': 'Low sugar',
    'Caloriearm': 'Low calorie',
    'Suikervrij': 'Low sugar',
    'Geen suiker ': 'Low sugar',
    'Diabeet': 'Diabetic',
    'Schaaldieren': 'No shellfish',
    'Gezonde , zo weinig mogelijk bewerkte voeding ': 'Little to no processed food',
    'Vegetarisch (geen vlees en vis),Vegan (geen dierlijke producten, inclusief zuivel en eieren)': 'Vegan',
    'fodmap': 'FODMAP',
    'Fodmap': 'FODMAP',
    'Vegetarisch (geen vlees en vis),Lactose intolerant (ik kan geen producten met lactose consumeren)': 'Vegatarian & Lactose intolerant',
    'Ik eet alles met mate om mijn gewicht onder controle te houden': 'Weight watching',
    'Lactose intolerant (ik kan geen producten met lactose consumeren),Pescotariër (vegetariër, maar ik eet vis/zeevruchten)': 'Lactose intolerant & Pescatarian',
    'Vegetarisch (geen vlees en vis),Flexitariër (vooral vegetarisch maar soms eet ik wel vlees/vis)': 'Flexitarian',
    'Pescotariër (vegetariër, maar ik eet vis/zeevruchten),Flexitariër (vooral vegetarisch maar soms eet ik wel vlees/vis)': 'Flexitarian',
    'lactose': 'Lactose intolerant',
    'Vegan (geen dierlijke producten, inclusief zuivel en eieren),Pescotariër (vegetariër, maar ik eet vis/zeevruchten)': 'Vegan & Pescatarian',
    'Voedingsallergieën': 'Allergies',
    'Koolhydraatarm': 'Low calorie',
    'Ik heb geen dieetbeperkingen,Lactose intolerant (ik kan geen producten met lactose consumeren),Pescotariër (vegetariër, maar ik eet vis/zeevruchten)': 'Lactose intolerant & Pescatarian',
    'weinig vet en suiker': 'Low fat & Low sugar',
    'Ik heb geen dieetbeperkingen,Vegan (geen dierlijke producten, inclusief zuivel en eieren)': 'Vegan',
    'Vetarm': 'Low fat',
    'Kha': 'Low calorie',
    'Intolérance au lactose (je ne peux pas consommer de produits contenant du lactose),Pescétarien (végétarien, mais je mange du poisson et les fruits de mer)': 'Lactose intolerant & Pescatarian',
    "Je n'ai aucune restriction alimentaire,Vegan (pas de produits d'origine animale, y compris les produits laitiers et les œufs)": 'Vegan',
    "Sans Gluten": 'Gluten free',
    "Je n'ai aucune restriction alimentaire,Autre, précisez...": 'None',
    "Végétarien (pas de viande ni de poisson),Vegan (pas de produits d'origine animale, y compris les produits laitiers et les œufs),Flexitarien (principalement végétarien, mais il m'arrive de manger de la viande ou du poisson)": 'Flexitarian',   
    'schaaldieren': 'No shellfish'
}

FL['Diet'] = FL['Diet'].replace(value_mapping)
WA['Diet'] = WA['Diet'].replace(value_mapping)

value_mapping = {
    "Je n'ai aucune restriction alimentaire": 'None',
    "Végétarien (pas de viande ni de poisson)": 'Vegetarian',
    "Vegan (pas de produits d'origine animale, y compris les produits laitiers et les œufs)": 'Vegan',
    "Intolérance au lactose (je ne peux pas consommer de produits contenant du lactose)": 'Lactose intolerant',
    "Pescétarien (végétarien, mais je mange du poisson et les fruits de mer)": 'Pescatarian',
    "Flexitarien (principalement végétarien, mais il m'arrive de manger de la viande ou du poisson)": 'Flexitarian',
    'sans sucre diabétique': 'Diabetic',
    'diabète': 'Diabetic',
    'allergies alimentaires': 'Allergies',
    "Végétarien (pas de viande ni de poisson),Pescétarien (végétarien, mais je mange du poisson et les fruits de mer)":'Pescatarian',
    'Vegetarisch (geen vlees en vis),Andere, namelijk ...': 'Vegetarian',
    'Quelques allergies foie gras fraises ': 'Allergies',
    'Peu de sucre': 'Low sugar',
    "Je n'ai aucune restriction alimentaire,Vegan (pas de produits d'origine animale, y compris les produits laitiers et les œufs),Pescétarien (végétarien, mais je mange du poisson et les fruits de mer),Flexitarien (principalement végétarien, mais il m'arrive de manger de la viande ou du poisson)": 'Flexitarian',
    "Intolérance au lactose (je ne peux pas consommer de produits contenant du lactose),Flexitarien (principalement végétarien, mais il m'arrive de manger de la viande ou du poisson)": 'Lactose intolerant & Flexitarian',
    "Intolérance au lactose (je ne peux pas consommer de produits contenant du lactose),Autre, précisez..."    : "Lactose intolerant",
    "Allergie à la protéine de lait " : 'Allergies',
    "Allergie aux cacahuète " : 'Allergies',
    "geen beperkingen": 'None',
    "diabeet": 'Diabetic',
    "diabetes": 'Diabetic',
    "Suiker arm": "Low sugar",
    "Halal ": 'Halal',
    "Je n'ai aucune restriction alimentaire,Flexitarien (principalement végétarien, mais il m'arrive de manger de la viande ou du poisson)": "Flexitarian",
    "Végétarien (pas de viande ni de poisson),Vegan (pas de produits d'origine animale, y compris les produits laitiers et les œufs),Intolérance au lactose (je ne peux pas consommer de produits contenant du lactose),Pescétarien (végétarien, mais je mange du poisson et les fruits de mer)" : "Pescatarian & Lactose intolerant"
}

FL['Diet'] = FL['Diet'].replace(value_mapping)
WA['Diet'] = WA['Diet'].replace(value_mapping)

FL.loc[FL['ID'] == '2,1166', 'Diet'] = 'Diabetic'
WA.loc[WA['ID'] == '1,973', 'Diet'] = 'Vegetarian & Low sugar'
WA.loc[WA['ID'] == '2,334', 'Diet'] = 'No porc'
WA.loc[WA['ID'] == '2,701', 'Diet'] = 'Lactose intolerant & No gluten'
WA.loc[WA['ID'] == '2,565', 'Diet'] = 'Lactose intolerant & Allergies'
WA.loc[WA['ID'] == '2,1375', 'Diet'] = 'Lactose intolerant & No gluten'


print(FL['Diet'].unique())
print(WA['Diet'].unique())

WA.drop(columns=['Q138_7_TEXT'], inplace=True)
FL.drop(columns=['Q138_7_TEXT'], inplace=True)

total = pd.concat([FL, WA], ignore_index=True)
print(total['Diet'].value_counts(normalize=True) * 100)

del value_mapping
del total

""" Adding an eco diet column to DCE """

def determine_eco_diet(row):
    if row['Region'] == 1:
        diet = FL.loc[FL['ID'] == row['ID'], 'Diet'].values[0]
    else:
        diet = WA.loc[WA['ID'] == row['ID'], 'Diet'].values[0]

    if "Vegan" in diet or "Vegetarian" in diet or "Pescatarian" in diet or "Flexitarian" in diet:
        return 1
    else:
        return 0

DCE['Eco diet'] = DCE.apply(determine_eco_diet, axis=1) # 12.10%

#%% Understanding the column Vraag 1a - Adaptability

""" Renaming the column """

FL.rename(columns={'Vraag 1a': 'Adaptability'}, inplace=True)
WA.rename(columns={'Vraag 1a': 'Adaptability'}, inplace=True)


""" Uniform values """

value_mapping = {
    'Soms': 'Sometimes',
    'Vaak': 'Often',
    'Nooit': 'Never',
    'Altijd': 'Always',
    'Souvent':'Often',
    'Jamais': 'Never',
    'Toujours': 'Always',
    'Parfois': 'Sometimes'
}

FL['Adaptability'] = FL['Adaptability'].replace(value_mapping)
WA['Adaptability'] = WA['Adaptability'].replace(value_mapping)

""" Checking frequencies """

combined_df = pd.concat([FL, WA], ignore_index=True)
adaptability_freq = combined_df['Adaptability'].value_counts(normalize=True) * 100  
print("Relative frequencies of unique values in the 'Adaptability' column:")
print(adaptability_freq)

""" Coding it as a continuous variable """

value_mapping = {
    'Never': 1,
    'Sometimes': 2,
    'Often': 3,
    'Always': 4
}

FL['Adaptability'] = FL['Adaptability'].replace(value_mapping)
WA['Adaptability'] = WA['Adaptability'].replace(value_mapping)


""" Adding apatability column to DCE """

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Adaptability'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Adaptability'].values[0]

    return x

DCE['Adaptability'] = DCE.apply(determine, axis=1)

#%% Understanding the column Inclusion 4  - Storage

""" Renaming the column """

FL.rename(columns={'Inclusion 4': 'Storage'}, inplace=True)
WA.rename(columns={'Inclusion 4': 'Storage'}, inplace=True)

""" Storage_1: when opened or unopened  """

def check_condition(value):
    if pd.isna(value):
        return 0
    if 'Ja, zowel geopend als ongeopend' in value or 'Oui, que le paquet soit ouvert ou non' in value:
        return 1
    return 0

FL.insert(3, 'Storage_1', FL['Storage'].apply(lambda x: check_condition(x)))
WA.insert(3, 'Storage_1', WA['Storage'].apply(lambda x: check_condition(x)))

""" Storage_2: opened """

def check_condition(value):
    if pd.isna(value):
        return 0
    if  'Ja, maar alleen wanneer het al geopend is' in value or 'Oui, mais seulement lorsque le paquet est  ouvert' in value:
        return 1
    return 0

FL.insert(4, 'Storage_2', FL['Storage'].apply(lambda x: check_condition(x)))
WA.insert(4, 'Storage_2', WA['Storage'].apply(lambda x: check_condition(x)))

""" Checking frequencies """

combined_df = pd.concat([FL, WA], ignore_index=True)
print(combined_df['Storage_1'].value_counts(normalize=True) * 100)  
print(combined_df['Storage_2'].value_counts(normalize=True) * 100) 


""" Deleting columns """

FL.drop(columns=['Storage'], inplace=True)
WA.drop(columns=['Storage'], inplace=True)

FL.drop(columns=['Storage_1'], inplace=True)
WA.drop(columns=['Storage_1'], inplace=True)

""" Adding apatability column to DCE """

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Storage_2'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Storage_2'].values[0]

    return x

DCE['Storage_2'] = DCE.apply(determine, axis=1)

#%% Understanding the column Q139  - Who shops for the household

""" Renaming the column """

FL.rename(columns={'Q139': 'Shopper'}, inplace=True)
WA.rename(columns={'Q139': 'Shopper'}, inplace=True)

""" Uniform values """

value_mapping = {
    'Altijd': 'Always',
    'Jamais': 'Never',
    'Nooit': 'Never',
    'Parfois': 'Sometimes',
    'Soms': 'Sometimes',
    'Souvent': 'Often',
    'Vaak': 'Often',
    'Toujours':'Always'
}

FL['Shopper'] = FL['Shopper'].replace(value_mapping)
WA['Shopper'] = WA['Shopper'].replace(value_mapping)


""" Checking frequencies """

combined_df = pd.concat([FL, WA], ignore_index=True)
freq = combined_df['Shopper'].value_counts(normalize=True) * 100  
print("Relative frequencies of unique values in the 'Shopper' column:")
print(freq)

""" Coding it as a continuous variable """

value_mapping = {
    'Never': 1,
    'Sometimes': 2,
    'Often': 3,
    'Always': 4
}

FL['Shopper'] = FL['Shopper'].replace(value_mapping)
WA['Shopper'] = WA['Shopper'].replace(value_mapping)

""" Adding shopper column to DCE """

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Shopper'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Shopper'].values[0]

    return x

DCE['Shopper'] = DCE.apply(determine, axis=1)

#%% Understanding the column Vraag 2a  - Checks good products at home before shopping

""" Renaming the column """

FL.rename(columns={'Vraag 2a ': 'Checker'}, inplace=True)
WA.rename(columns={'Vraag 2a ': 'Checker'}, inplace=True)

""" Uniform values """

value_mapping = {
    'Altijd': 'Always',
    'Jamais': 'Never',
    'Nooit': 'Never',
    'Parfois': 'Sometimes',
    'Soms': 'Sometimes',
    'Souvent': 'Often',
    'Vaak': 'Often',
    'Toujours':'Always'
}

FL['Checker'] = FL['Checker'].replace(value_mapping)
WA['Checker'] = WA['Checker'].replace(value_mapping)

FL['Checker'] = FL['Checker'].fillna('Unknown')
WA['Checker'] = WA['Checker'].fillna('Unknown')

""" Checking frequencies """

combined_df = pd.concat([FL, WA], ignore_index=True)
freq = combined_df['Checker'].value_counts(normalize=True) * 100  
print("Relative frequencies of unique values in the 'Checker' column:")
print(freq)

""" Coding it as a continuous variable """

value_mapping = {
    'Never': 1,
    'Sometimes': 2,
    'Often': 3,
    'Always': 4,
    'Unknown': 0
}

FL['Checker'] = FL['Checker'].replace(value_mapping)
WA['Checker'] = WA['Checker'].replace(value_mapping)

""" Adding Checker column to DCE """

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Checker'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Checker'].values[0]

    return x

DCE['Checker'] = DCE.apply(determine, axis=1)


#%% Understanding the column Vraag 2b  - WHY? Checks good products at home before shopping

""" Checker_1: to check necessary products  """

def check_condition(value):
    if pd.isna(value):
        return 0
    if 'Om te controleren welke voedingsproducten er nodig zijn uit de winkel' in value or 'Pour vérifier les produits alimentaires qui manquent' in value:
        return 1
    return 0

FL.insert(16, 'Checker_1', FL['Vraag 2b '].apply(lambda x: check_condition(x)))
WA.insert(16, 'Checker_1', WA['Vraag 2b '].apply(lambda x: check_condition(x)))

""" Checker_2: to make place for new products """

def check_condition(value):
    if pd.isna(value):
        return 0
    if  'Om ruimte te maken voor nieuwe boodschappen' in value or 'Pour faire de la place aux nouveaux produits achetés' in value:
        return 1
    return 0

FL.insert(17, 'Checker_2', FL['Vraag 2b '].apply(lambda x: check_condition(x)))
WA.insert(17, 'Checker_2', WA['Vraag 2b '].apply(lambda x: check_condition(x)))

""" Checker_3: to throw away spoiled food items """

def check_condition(value):
    if pd.isna(value):
        return 0
    if  'Om mogelijk bedorven voedsel weg te gooien' in value or 'Pour jeter les aliments potentiellement avariés' in value:
        return 1
    return 0

FL.insert(18, 'Checker_3', FL['Vraag 2b '].apply(lambda x: check_condition(x)))
WA.insert(18, 'Checker_3', WA['Vraag 2b '].apply(lambda x: check_condition(x)))

""" Vraag 2b_4_TEXT """

# FL
# 1,1423: verspilling vermijden
# 1,277: om te controleren wat er eerst wordt gegetenin verband met bederf


# WA

# 2,380: Pour ne pas acheter si pas besoin et éviter ainsi le gaspillage ==> 1, but already selected 1
# 2,436: Pour éviter le gaspillage 
# 2,198: afin d'éviter le gaspillage en utilisant ce qui me reste à la maison
# 2,767: pour les cuisiner avec d'autres produits
# 2,243: pour ne pas gaspillé 
# 2,402: pour ne rien oubler ==> 1, but already selected 1

""" Checking frequencies """

combined_df = pd.concat([FL, WA], ignore_index=True)
print(combined_df['Checker_1'].value_counts(normalize=True) * 100)  
print(combined_df['Checker_2'].value_counts(normalize=True) * 100)  
print(combined_df['Checker_3'].value_counts(normalize=True) * 100)  

""" Deleting columns """

FL.drop(columns=['Vraag 2b '], inplace=True)
WA.drop(columns=['Vraag 2b '], inplace=True)

FL.drop(columns=['Vraag 2b _4_TEXT'], inplace=True)
WA.drop(columns=['Vraag 2b _4_TEXT'], inplace=True)

""" Adding Checker_1, Checker_2 and Checker_3 column to DCE """

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Checker_1'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Checker_1'].values[0]

    return x

DCE['Checker_1'] = DCE.apply(determine, axis=1)

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Checker_2'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Checker_2'].values[0]

    return x

DCE['Checker_2'] = DCE.apply(determine, axis=1)

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Checker_3'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Checker_3'].values[0]

    return x

DCE['Checker_3'] = DCE.apply(determine, axis=1)
#%% Understanding the column Vraag 4a  - Buy discounted food products

""" Renaming the column """

FL.rename(columns={'Vraag 4a ': 'Discounted'}, inplace=True)
WA.rename(columns={'Vraag 4a ': 'Discounted'}, inplace=True)

""" Uniform values """

value_mapping = {
    'Altijd': 'Always',
    'Jamais': 'Never',
    'Nooit': 'Never',
    'Parfois': 'Sometimes',
    'Soms': 'Sometimes',
    'Souvent': 'Often',
    'Vaak': 'Often',
    'Toujours':'Always'
}

FL['Discounted'] = FL['Discounted'].replace(value_mapping)
WA['Discounted'] = WA['Discounted'].replace(value_mapping)

FL['Discounted'] = FL['Discounted'].fillna('Unknown')
WA['Discounted'] = WA['Discounted'].fillna('Unknown')

""" Checking frequencies """

combined_df = pd.concat([FL, WA], ignore_index=True)
freq = combined_df['Discounted'].value_counts(normalize=True) * 100  
print("Relative frequencies of unique values in the 'Discounted' column:")
print(freq)

""" Coding it as a continuous variable """

value_mapping = {
    'Never': 1,
    'Sometimes': 2,
    'Often': 3,
    'Always': 4,
    'Unknown': 0
}

FL['Discounted'] = FL['Discounted'].replace(value_mapping)
WA['Discounted'] = WA['Discounted'].replace(value_mapping)

""" Adding Discounted column to DCE """


def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Discounted'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Discounted'].values[0]

    return x

DCE['Discounted'] = DCE.apply(determine, axis=1)

#%% Understanding the column Vraag 4b  - WHY? buy discounted food items when they are about to spoil

""" Discounted_1: economic advantage  """

def check_condition(value):
    if pd.isna(value):
        return 0
    if 'Vanwege het economische voordeel' in value or 'Pour faire des économies' in value:
        return 1
    return 0

FL.insert(20, 'Discounted_1', FL['Vraag 4b'].apply(lambda x: check_condition(x)))
WA.insert(20, 'Discounted_1', WA['Vraag 4b'].apply(lambda x: check_condition(x)))

""" Discounted_2: eco-conscious """

def check_condition(value):
    if pd.isna(value):
        return 0
    if  'Dit is een milieubewuste keuze' in value or "Il s'agit d'un choix respectueux de l'environnement" in value:
        return 1
    return 0

FL.insert(21, 'Discounted_2', FL['Vraag 4b'].apply(lambda x: check_condition(x)))
WA.insert(21, 'Discounted_2', WA['Vraag 4b'].apply(lambda x: check_condition(x)))

""" Discounted_3: trying new products 1"""

def check_condition(value):
    if pd.isna(value):
        return 0
    if  'Om voedingsproducten te proberen die ik normaal niet zou kopen' in value or "Pour essayer de nouveaux produits alimentaires que je n'achèterais pas normalement" in value:
        return 1
    return 0

FL.insert(22, 'Discounted_3', FL['Vraag 4b'].apply(lambda x: check_condition(x)))
WA.insert(22, 'Discounted_3', WA['Vraag 4b'].apply(lambda x: check_condition(x)))

""" Other reasons """

# FL
# 1,589: Als ik deze producten tocht de dag zelf nog nodig heb
# 1,761: Als ik ze kan invriezen 
# 1,699: Als ik ze onmiddellijk kan consumeren
# 1,440: Goedkoper ==> 1
FL.loc[FL['ID'] == '1,440', 'Discounted_1'] = 1
# 1,437: Prijs  ==> 1, but already selected 1
# 1,1337: om in te vriezen en een andere keer te eten of gebruiken
# 1,1085: omdat ik die juist nodig heb
# 1,305: prijs ==> 1
FL.loc[FL['ID'] == '1,305', 'Discounted_1'] = 1
# 1,1041: voedselverspilling ==> 2
FL.loc[FL['ID'] == '1,1041', 'Discounted_2'] = 1
# 1,250: waarom niet als ik ze dezelfde avond verwerkt
# 2,44: Sans raison autre que le produit 
# 2,1145: ik begrijp de vraag net

# WA
# 2,869: Si je prévois de le manger le jour même 
# 2,590: quand je sais que je vais les consommer rapidement
# 2,174: sinon ça vas a la poubelle "des animaux ont donner leur vie pour nous nourrir" 

""" Checking frequencies """

combined_df = pd.concat([FL, WA], ignore_index=True)
print(combined_df['Discounted_1'].value_counts(normalize=True) * 100)  
print(combined_df['Discounted_2'].value_counts(normalize=True) * 100)  
print(combined_df['Discounted_3'].value_counts(normalize=True) * 100)  

""" Deleting columns """

FL.drop(columns=['Vraag 4b'], inplace=True)
WA.drop(columns=['Vraag 4b'], inplace=True)

FL.drop(columns=['Vraag 4b_4_TEXT'], inplace=True)
WA.drop(columns=['Vraag 4b_4_TEXT'], inplace=True)

""" Adding Discounted_1, Discounted_2 and Discounted_3 column to DCE """

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Discounted_1'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Discounted_1'].values[0]

    return x

DCE['Discounted_1'] = DCE.apply(determine, axis=1)

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Discounted_2'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Discounted_2'].values[0]

    return x

DCE['Discounted_2'] = DCE.apply(determine, axis=1)

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Discounted_3'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Discounted_3'].values[0]

    return x

DCE['Discounted_3'] = DCE.apply(determine, axis=1)


#%% Understanding the column Vraag 5a  - Make use of services

""" Renaming the column """

FL.rename(columns={'Vraag 5a ': 'Services'}, inplace=True)
WA.rename(columns={'Vraag 5a ': 'Services'}, inplace=True)

""" Uniform values """

value_mapping = {
    'Altijd': 'Always',
    'Jamais': 'Never',
    'Nooit': 'Never',
    'Parfois': 'Sometimes',
    'Soms': 'Sometimes',
    'Souvent': 'Often',
    'Vaak': 'Often',
    'Toujours':'Always'
}

FL['Services'] = FL['Services'].replace(value_mapping)
WA['Services'] = WA['Services'].replace(value_mapping)

FL['Services'] = FL['Services'].fillna('Unknown')
WA['Services'] = WA['Services'].fillna('Unknown')

""" Checking frequencies """

combined_df = pd.concat([FL, WA], ignore_index=True)
freq = combined_df['Services'].value_counts(normalize=True) * 100  
print("Relative frequencies of unique values in the 'Services' column:")
print(freq)

""" Coding it as a continuous variable """

value_mapping = {
    'Never': 1,
    'Sometimes': 2,
    'Often': 3,
    'Always': 4,
    'Unknown': 0
}

FL['Services'] = FL['Services'].replace(value_mapping)
WA['Services'] = WA['Services'].replace(value_mapping)

""" Adding Services column to DCE """

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Services'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Services'].values[0]

    return x

DCE['Services'] = DCE.apply(determine, axis=1)

#%% Understanding the column Vraag 5c  - WHY Make use of services


""" Services_1: economic advantage  """

def check_condition(value):
    if pd.isna(value):
        return 0
    if 'Vanwege het economische voordeel' in value or 'Pour faire des économies' in value:
        return 1
    return 0

FL.insert(24, 'Services_1', FL['Vraag 5c '].apply(lambda x: check_condition(x)))
WA.insert(24, 'Services_1', WA['Vraag 5c '].apply(lambda x: check_condition(x)))

""" Discounted_2: eco-conscious """

def check_condition(value):
    if pd.isna(value):
        return 0
    if  'Dit is een milieubewuste keuze' in value or "Il s'agit d'un choix respectueux de l'environnement" in value:
        return 1
    return 0

FL.insert(25, 'Services_2', FL['Vraag 5c '].apply(lambda x: check_condition(x)))
WA.insert(25, 'Services_2', WA['Vraag 5c '].apply(lambda x: check_condition(x)))

""" Discounted_3: trying new products 1"""

def check_condition(value):
    if pd.isna(value):
        return 0
    if  'Om producten te proberen die ik normaal niet zou kopen' in value or "Pour essayer de nouveaux produits que je n'achèterais pas normalement" in value:
        return 1
    return 0

FL.insert(26, 'Services_3', FL['Vraag 5c '].apply(lambda x: check_condition(x)))
WA.insert(26, 'Services_3', WA['Vraag 5c '].apply(lambda x: check_condition(x)))

""" Other reasons """

# FL
# 1,585: Om andere mensen te helpen
# 1,1146: Verrassingselement ==> 3
# 1,1337: de vriezer aan te vullen met vlees of groenten
# 1,653: financiële redenen ==> 1
# 1,1041: voedselverspilling
# 2,1000: le restaurant de ma societe vend des restes du repas chaud de midi a des prix derisoire. et tjs tres frais

FL.loc[FL['ID'] == '1,653', 'Services_1'] = 1

# WA
# 2,97: pour la surprise
# 2,330: jeu

""" Checking frequencies """

combined_df = pd.concat([FL, WA], ignore_index=True)
print(combined_df['Services_1'].value_counts(normalize=True) * 100)  
print(combined_df['Services_2'].value_counts(normalize=True) * 100)  
print(combined_df['Services_3'].value_counts(normalize=True) * 100)  

""" Deleting columns """

FL.drop(columns=['Vraag 5c '], inplace=True)
WA.drop(columns=['Vraag 5c '], inplace=True)

FL.drop(columns=['Vraag 5c _4_TEXT'], inplace=True)
WA.drop(columns=['Vraag 5c _4_TEXT'], inplace=True)

""" Adding Services_1, Services_2 and Services_3 column to DCE """

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Services_1'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Services_1'].values[0]

    return x

DCE['Services_1'] = DCE.apply(determine, axis=1)

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Services_2'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Services_2'].values[0]

    return x

DCE['Services_2'] = DCE.apply(determine, axis=1)

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Services_3'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Services_3'].values[0]

    return x

DCE['Services_3'] = DCE.apply(determine, axis=1)

#%% Understanding the column Vraag 6  - Where discarding food items


""" Disposal_1: residual waste bag  """

def check_condition(value):
    if pd.isna(value):
        return 0
    if 'Ik gooi het in de restafvalzak' in value or 'Je les jette dans le sac à déchets résiduels' in value:
        return 1
    return 0

FL.insert(27, 'Disposal_1', FL['Vraag 6'].apply(lambda x: check_condition(x)))
WA.insert(27, 'Disposal_1', WA['Vraag 6'].apply(lambda x: check_condition(x)))

""" Disposal_2: the separate garbage bag (for vegetables, fruit and garden waste) """

def check_condition(value):
    if pd.isna(value):
        return 0
    if  'Ik gooi het in de aparte groente-, fruit-, en tuinafval (GFT) vuilniszak' in value or "Je les jette dans le sac poubelle séparé (pour les légumes, les fruits et les déchets de jardin)" in value:
        return 1
    return 0

FL.insert(28, 'Disposal_2', FL['Vraag 6'].apply(lambda x: check_condition(x)))
WA.insert(28, 'Disposal_2', WA['Vraag 6'].apply(lambda x: check_condition(x)))

""" Disposal_3: pets"""

def check_condition(value):
    if pd.isna(value):
        return 0
    if  'Ik geef het aan (huis)dieren' in value or "Je les donne aux animaux domestiques" in value:
        return 1
    return 0

FL.insert(29, 'Disposal_3', FL['Vraag 6'].apply(lambda x: check_condition(x)))
WA.insert(29, 'Disposal_3', WA['Vraag 6'].apply(lambda x: check_condition(x)))

""" Disposal_4: composting """

def check_condition(value):
    if pd.isna(value):
        return 0
    if  'Ik composteer het' in value or "Je les composte" in value:
        return 1
    return 0

FL.insert(30, 'Disposal_4', FL['Vraag 6'].apply(lambda x: check_condition(x)))
WA.insert(30, 'Disposal_4', WA['Vraag 6'].apply(lambda x: check_condition(x)))

""" Other reasons """

# FL
# 1,273: Aan de hond --> 3
# 1,588: Ik gooi het vanachter in de tuin op een hoop --> 4
# 1,1054: Ik heb nooit voedselafval
# 1,603: Ik heb praktisch nooit restafval
# 1,1218: Kippen --> 3
# 1,546: geen
# 1,602: groene container --> 2
# 1,852: ik gooi niks weg
# 1,262: ik gppi geen voedselafvazl weg of zo weinig mogelijk
# 1,1493: ik heb geen vodselafval
# 1,893: mesthoop --> 4
# 2,1130: Ik geef dit aan de kippen en de rest bij de gft bak ==> 3

FL.loc[FL['ID'] == '1,893', 'Disposal_4'] = 1
FL.loc[FL['ID'] == '1,602', 'Disposal_2'] = 1
FL.loc[FL['ID'] == '1,588', 'Disposal_4'] = 1
FL.loc[FL['ID'] == '1,1218', 'Disposal_3'] = 1
FL.loc[FL['ID'] == '1,273', 'Disposal_3'] = 1
FL.loc[FL['ID'] == '2,1130', 'Disposal_3'] = 1


# WA
# 2,854: Compost -- 3
# 2,482: Je ne jeté jamais de reste
# 2,114: Nous avons peu de déchets car j'achète que ce dont j'ai besoin
# 2,207: Poules ==> 3
# 2,447: Sac orange
# 2,209: chez le fermier
# 2,701: compost ==> 4
# 2,97: compostage ==> 4
# 2,313: dans ma poubelle normale ==> 2
# 2,370: j'ai des poules ==> 3
# 2,198: je les jette dans des bulles, poubelles  à cet effet ==> 3
# 2,842: les donne pour les poules  ==> 3
# 2,590: sur le compost du jardin ==> 4

WA.loc[WA['ID'] == '2,854', 'Disposal_3'] = 1
WA.loc[WA['ID'] == '2,207', 'Disposal_3'] = 1
WA.loc[WA['ID'] == '2,701', 'Disposal_4'] = 1
WA.loc[WA['ID'] == '2,97', 'Disposal_4'] = 1
WA.loc[WA['ID'] == '2,313', 'Disposal_2'] = 1
WA.loc[WA['ID'] == '2,370', 'Disposal_3'] = 1
WA.loc[WA['ID'] == '2,198', 'Disposal_3'] = 1
WA.loc[WA['ID'] == '2,842', 'Disposal_3'] = 1
WA.loc[WA['ID'] == '2,590', 'Disposal_4'] = 1

""" Checking frequencies """

combined_df = pd.concat([FL, WA], ignore_index=True)
print(combined_df['Disposal_1'].value_counts(normalize=True) * 100)  
print(combined_df['Disposal_2'].value_counts(normalize=True) * 100)  
print(combined_df['Disposal_3'].value_counts(normalize=True) * 100)  
print(combined_df['Disposal_4'].value_counts(normalize=True) * 100)  


""" Deleting columns """

FL.drop(columns=['Vraag 6'], inplace=True)
WA.drop(columns=['Vraag 6'], inplace=True)

FL.drop(columns=['Vraag 6_5_TEXT'], inplace=True)
WA.drop(columns=['Vraag 6_5_TEXT'], inplace=True)


""" Adding Disposal_1 column to DCE """

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Disposal_1'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Disposal_1'].values[0]

    return x

DCE['Disposal_1'] = DCE.apply(determine, axis=1)

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Disposal_2'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Disposal_2'].values[0]

    return x

DCE['Disposal_2'] = DCE.apply(determine, axis=1)

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Disposal_3'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Disposal_3'].values[0]

    return x

DCE['Disposal_3'] = DCE.apply(determine, axis=1)

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Disposal_4'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Disposal_4'].values[0]

    return x

DCE['Disposal_4'] = DCE.apply(determine, axis=1)

#%% Understanding the column Vraag 7  - What do you do with leftovers?


""" Leftovers_1: throw them away  """

def check_condition(value):
    if pd.isna(value):
        return 0
    if 'Ik gooi deze weg' in value or 'Je les jette' in value:
        return 1
    return 0

FL.insert(31, 'Leftovers_1', FL['Vraag 7'].apply(lambda x: check_condition(x)))
WA.insert(31, 'Leftovers_1', WA['Vraag 7'].apply(lambda x: check_condition(x)))

""" Leftovers_2: in the fridge """

def check_condition(value):
    if pd.isna(value):
        return 0
    if  'Ik bewaar deze in de koelkast voor één van de volgende dagen' in value or "Je les conserve au réfrigérateur pour les jours suivants" in value:
        return 1
    return 0

FL.insert(32, 'Leftovers_2', FL['Vraag 7'].apply(lambda x: check_condition(x)))
WA.insert(32, 'Leftovers_2', WA['Vraag 7'].apply(lambda x: check_condition(x)))

""" Leftovers_3: in the freezer"""

def check_condition(value):
    if pd.isna(value):
        return 0
    if  'Ik bewaar deze in de vriezer voor een andere keer' in value or "Je les garde au congélateur pour une autre fois" in value:
        return 1
    return 0

FL.insert(33, 'Leftovers_3', FL['Vraag 7'].apply(lambda x: check_condition(x)))
WA.insert(33, 'Leftovers_3', WA['Vraag 7'].apply(lambda x: check_condition(x)))

""" Disposal_4: give them away """

def check_condition(value):
    if pd.isna(value):
        return 0
    if  'Ik geef deze aan een (on)bekende die het kan gebruiken' in value or "Je les offre à quelqu'un qui pourra les utiliser" in value:
        return 1
    return 0

FL.insert(34, 'Leftovers_4', FL['Vraag 7'].apply(lambda x: check_condition(x)))
WA.insert(34, 'Leftovers_4', WA['Vraag 7'].apply(lambda x: check_condition(x)))

""" Other reasons """

# FL
#1,1317: De hond krijgt het vaak ==> 4
FL.loc[FL['ID'] == '1,1317', 'Leftovers_4'] = 1
#1,1248: De hond krijgt ook restjes ==> 4
FL.loc[FL['ID'] == '1,1248', 'Leftovers_4'] = 1
#1,752: Hond ===> 4
FL.loc[FL['ID'] == '1,752', 'Leftovers_4'] = 1
#1,1068: Hond of kippen ==> 4
FL.loc[FL['ID'] == '1,1068', 'Leftovers_4'] = 1
#1,1054: Ik gebruik deze steeds ==> 2, 3
# 2,942: Ik neem ze mee als lunchpakket voor de volgende dag.
#1,740: Ik geef het aan de kippen ==> 4
FL.loc[FL['ID'] == '1,640', 'Leftovers_4'] = 1
#1,1423: Ik maak er soep van ==> 2,3
# 2,1082: Meestal eet ik het later die avond op
#1,316: Vogels ==> 4
FL.loc[FL['ID'] == '1,316', 'Leftovers_4'] = 1
#1,233: Voor de kippen ==> 4
FL.loc[FL['ID'] == '1,233', 'Leftovers_4'] = 1
#1,140: We hebben  normaal nooit restjes, in de uitzonderlijke gevallen zouden deze voor de kippen zijn ==> 4
FL.loc[FL['ID'] == '1,14O', 'Leftovers_4'] = 1
#1,1254: Wij hebben nooit restjes. Ik weeg alles af
#1,791: aan onze hond ==> 4
FL.loc[FL['ID'] == '1,791', 'Leftovers_4'] = 1
#1,893: er zijn geen restjes (afmeten ingredienten)
#1,596: er zijn nooit restjes
#1,349: hond ==> 4
FL.loc[FL['ID'] == '1,349', 'Leftovers_4'] = 1
#1,250: ik bel de buren op ,als ze willen komen ze het ophalen ==> 4
FL.loc[FL['ID'] == '1,250', 'Leftovers_4'] = 1
#1,940: ik maak er soep van ==> 2,3
FL.loc[FL['ID'] == '1,940', 'Leftovers_2'] = 1
FL.loc[FL['ID'] == '1,940', 'Leftovers_3'] = 1
#1,876: kippen ==> 4
FL.loc[FL['ID'] == '1,876', 'Leftovers_4'] = 1
#1,267: kleine hapjes voor de  kippen ==> 4
FL.loc[FL['ID'] == '1,276', 'Leftovers_4'] = 1
#1,889: nooit restjes over
#1,237: soms eet ik het de volgende dag op en soms gooi ik het weg ==> 1,2
FL.loc[FL['ID'] == '1,237', 'Leftovers_2'] = 1
FL.loc[FL['ID'] == '1,237', 'Leftovers_1'] = 1

# WA
# 2,486: En général il n'y a pas de reste
# 2,55: Il reste rarement 
# 2,851: Je donne au chien et chat ==> 4
WA.loc[WA['ID'] == '2,851', 'Leftovers_4'] = 1
# 2,165: Ou je vais les donner au sdf dans la rue  ==> 4
WA.loc[WA['ID'] == '2,165', 'Leftovers_4'] = 1
# 2,339: Pour le chien ==> 4
WA.loc[WA['ID'] == '2,339', 'Leftovers_4'] = 1
# 2,1099: Pour le chien  ==> 4
WA.loc[WA['ID'] == '2,1099', 'Leftovers_4'] = 1
# 2,97: animaux ==> 4
WA.loc[WA['ID'] == '2,97', 'Leftovers_4'] = 1
# 2,351: compost ==> 1
WA.loc[WA['ID'] == '2,351', 'Leftovers_1'] = 1
# 2,209: dans plusieurs seau puis va chez le fermier
# 2,47: il n`y en a pas
# 2,310: je les donne à mes poules ==> 4
WA.loc[WA['ID'] == '2,310', 'Leftovers_4'] = 1

""" Checking frequencies """

combined_df = pd.concat([FL, WA], ignore_index=True)
print(combined_df['Leftovers_1'].value_counts(normalize=True) * 100)  
print(combined_df['Leftovers_2'].value_counts(normalize=True) * 100)  
print(combined_df['Leftovers_3'].value_counts(normalize=True) * 100)  
print(combined_df['Leftovers_4'].value_counts(normalize=True) * 100)  


""" Deleting columns """

FL.drop(columns=['Vraag 7'], inplace=True)
WA.drop(columns=['Vraag 7'], inplace=True)

FL.drop(columns=['Vraag 7_5_TEXT'], inplace=True)
WA.drop(columns=['Vraag 7_5_TEXT'], inplace=True)

""" Adding Leftovers_1 column to DCE """

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Leftovers_1'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Leftovers_1'].values[0]

    return x

DCE['Leftovers_1'] = DCE.apply(determine, axis=1)

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Leftovers_2'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Leftovers_2'].values[0]

    return x

DCE['Leftovers_2'] = DCE.apply(determine, axis=1)

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Leftovers_3'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Leftovers_3'].values[0]

    return x

DCE['Leftovers_3'] = DCE.apply(determine, axis=1)

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Leftovers_4'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Leftovers_4'].values[0]

    return x

DCE['Leftovers_4'] = DCE.apply(determine, axis=1)

#%% Understanding the column Vraag Q85  - Understanding of the Best Before date label


""" BBproduct_1: meat and fish  """

def check_condition(value):
    if pd.isna(value):
        return 0
    if 'Vlees en vis' in value or 'Viande et poisson' in value:
        return 1
    return 0

FL.insert(35, 'BBproduct_1', FL['Q85'].apply(lambda x: check_condition(x)))
WA.insert(35, 'BBproduct_1', WA['Q85'].apply(lambda x: check_condition(x)))

""" BBproduct_2: tomato sauce """

def check_condition(value):
    if pd.isna(value):
        return 0
    if  'Tomatensaus' in value or "Sauce tomate" in value:
        return 1
    return 0

FL.insert(36, 'BBproduct_2', FL['Q85'].apply(lambda x: check_condition(x)))
WA.insert(36, 'BBproduct_2', WA['Q85'].apply(lambda x: check_condition(x)))

""" BBproduct_3: cooking cream"""

def check_condition(value):
    if pd.isna(value):
        return 0
    if  'Kookroom' in value or "Crème à cuisson liquide" in value:
        return 1
    return 0

FL.insert(37, 'BBproduct_3', FL['Q85'].apply(lambda x: check_condition(x)))
WA.insert(37, 'BBproduct_3', WA['Q85'].apply(lambda x: check_condition(x)))

""" BBproduct_4: pasta """

def check_condition(value):
    if pd.isna(value):
        return 0
    if  'Pasta' in value or "Pâtes" in value:
        return 1
    return 0

FL.insert(38, 'BBproduct_4', FL['Q85'].apply(lambda x: check_condition(x)))
WA.insert(38, 'BBproduct_4', WA['Q85'].apply(lambda x: check_condition(x)))

""" BBproduct_5: milk """

def check_condition(value):
    if pd.isna(value):
        return 0
    if  'Melk' in value or "Lait" in value:
        return 1
    return 0

FL.insert(39, 'BBproduct_5', FL['Q85'].apply(lambda x: check_condition(x)))
WA.insert(39, 'BBproduct_5', WA['Q85'].apply(lambda x: check_condition(x)))

""" BBproduct_6: don't know """

def check_condition(value):
    if pd.isna(value):
        return 0
    if  'Ik heb geen idee' in value or "Je ne sais pas" in value:
        return 1
    return 0

FL.insert(40, 'BBproduct_6', FL['Q85'].apply(lambda x: check_condition(x)))
WA.insert(40, 'BBproduct_6', WA['Q85'].apply(lambda x: check_condition(x)))

""" Checking frequencies """

combined_df = pd.concat([FL, WA], ignore_index=True)
print(combined_df['BBproduct_1'].value_counts(normalize=True) * 100)  
print(combined_df['BBproduct_2'].value_counts(normalize=True) * 100)  
print(combined_df['BBproduct_3'].value_counts(normalize=True) * 100)  
print(combined_df['BBproduct_4'].value_counts(normalize=True) * 100)  
print(combined_df['BBproduct_5'].value_counts(normalize=True) * 100)  
print(combined_df['BBproduct_6'].value_counts(normalize=True) * 100)  

""" Creating the BBproduct column """

FL['BBproduct'] = (FL['BBproduct_1'] | FL['BBproduct_6']).astype(int)
WA['BBproduct'] = (WA['BBproduct_1'] | WA['BBproduct_6']).astype(int)

""" Deleting columns """

FL.drop(columns=['Q85'], inplace=True)
WA.drop(columns=['Q85'], inplace=True)
FL.drop(columns=['BBproduct_1'], inplace=True)
WA.drop(columns=['BBproduct_1'], inplace=True)
FL.drop(columns=['BBproduct_2'], inplace=True)
WA.drop(columns=['BBproduct_2'], inplace=True)
FL.drop(columns=['BBproduct_3'], inplace=True)
WA.drop(columns=['BBproduct_3'], inplace=True)
FL.drop(columns=['BBproduct_4'], inplace=True)
WA.drop(columns=['BBproduct_4'], inplace=True)
FL.drop(columns=['BBproduct_5'], inplace=True)
WA.drop(columns=['BBproduct_5'], inplace=True)
FL.drop(columns=['BBproduct_6'], inplace=True)
WA.drop(columns=['BBproduct_6'], inplace=True)


""" Adding the BBproduct columns to DCE """

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'BBproduct'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'BBproduct'].values[0]

    return x

DCE['BBproduct'] = DCE.apply(determine, axis=1)

#%% Checking column Vraag 20b and Vraag 20b _5_TEXT

""" Conjoining the two columns """

for index, row in FL.iterrows():
    if row['Vraag 20b '] in ['Andere, namelijk ...', 'Autre, précisez ...']:
        FL.at[index, 'Vraag 20b '] = row['Vraag 20b _5_TEXT']

for index, row in WA.iterrows():
    if row['Vraag 20b '] in ['Andere, namelijk ...', 'Autre, précisez ...']:
        WA.at[index, 'Vraag 20b '] = row['Vraag 20b _5_TEXT']
        
FL.drop(columns=['Vraag 20b _5_TEXT'], inplace=True)
WA.drop(columns=['Vraag 20b _5_TEXT'], inplace=True)

""" BBdiscard_1: throw away  """

def check_condition(value):
    if pd.isna(value):
        return 0
    if 'gooi ik het product weg, ook als het nog ongeopend is.' in value or "Je jette le produit, même s'il n'est pas encore ouvert" in value:
        return 1
    return 0

FL.insert(40, 'BBdiscard_1', FL['Vraag 20b '].apply(lambda x: check_condition(x)))
WA.insert(40, 'BBdiscard_1', WA['Vraag 20b '].apply(lambda x: check_condition(x)))

""" BBdiscard_2: throw away if opened """

def check_condition(value):
    if pd.isna(value):
        return 0
    if  'gooi ik het product weg als het geopend is, als het ongeopend is dan geef ik het nog een kans.' in value or "Je jette le produit s'il est ouvert, s'il n'est pas ouvert, je le garde" in value:
        return 1
    return 0

FL.insert(42, 'BBdiscard_2', FL['Vraag 20b '].apply(lambda x: check_condition(x)))
WA.insert(42, 'BBdiscard_2', WA['Vraag 20b '].apply(lambda x: check_condition(x)))

""" BBdiscard_3: the look smell and tast"""

def check_condition(value):
    if pd.isna(value):
        return 0
    if  'kijk, ruik en/of proef ik het product om te controleren of het nog goed is.' in value or "Je regarde, je sens et/ou je goûte le produit pour vérifier s'il est encore bon" in value:
        return 1
    return 0

FL.insert(43, 'BBdiscard_3', FL['Vraag 20b '].apply(lambda x: check_condition(x)))
WA.insert(43, 'BBdiscard_3', WA['Vraag 20b '].apply(lambda x: check_condition(x)))

""" BBdiscard_4: holds no relevance """

def check_condition(value):
    if pd.isna(value):
        return 0
    if  'maakt dat voor mij niet uit. Ik kijk, ruik en/of proef het product altijd om te controleren of het nog goed is.' in value or "La DDM n'a pas d'importance pour moi. Je regarde, je sens et/ou je goûte toujours le produit pour vérifier s'il est encore bon" in value:
        return 1
    return 0

FL.insert(44, 'BBdiscard_4', FL['Vraag 20b '].apply(lambda x: check_condition(x)))
WA.insert(44, 'BBdiscard_4', WA['Vraag 20b '].apply(lambda x: check_condition(x)))

""" Other reasons """

# Fl
# 1,551: Laat eerst de kat proberen of zij het nog eet. 
# 1,889: kijk naar de datum vooraleer ik het koop

""" Checking frequencies """

combined_df = pd.concat([FL, WA], ignore_index=True)
print(combined_df['BBdiscard_1'].value_counts(normalize=True) * 100)  
print(combined_df['BBdiscard_2'].value_counts(normalize=True) * 100)  
print(combined_df['BBdiscard_3'].value_counts(normalize=True) * 100)  
print(combined_df['BBdiscard_4'].value_counts(normalize=True) * 100)  

""" Deleting columns """

FL.drop(columns=['Vraag 20b '], inplace=True)
WA.drop(columns=['Vraag 20b '], inplace=True)

""" Adding the BBdiscard columns to DCE """

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'BBdiscard_1'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'BBdiscard_1'].values[0]

    return x

DCE['BBdiscard_1'] = DCE.apply(determine, axis=1)

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'BBdiscard_2'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'BBdiscard_2'].values[0]

    return x

DCE['BBdiscard_2'] = DCE.apply(determine, axis=1)

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'BBdiscard_3'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'BBdiscard_3'].values[0]

    return x

DCE['BBdiscard_3'] = DCE.apply(determine, axis=1)

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'BBdiscard_4'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'BBdiscard_4'].values[0]

    return x

DCE['BBdiscard_4'] = DCE.apply(determine, axis=1)

#%% Understanding the column Q170  - Imagining of product origine: animal- or plant based 


""" Origine_1: plant-based  """

def check_condition(value):
    if pd.isna(value):
        return 0
    if 'Ik heb me voorgesteld dat het plantaardig was (zoals soja of kokos).' in value or 'Une crème à base végétale (ex: soja, noix de coco, ...)' in value:
        return 1
    return 0

FL.insert(45, 'Origine_1', FL['Q170'].apply(lambda x: check_condition(x)))
WA.insert(45, 'Origine_1', WA['Q170'].apply(lambda x: check_condition(x)))

""" Origine_2: animal based """

def check_condition(value):
    if pd.isna(value):
        return 0
    if  'Ik heb me voorgesteld dat het op dierlijke basis was (zuivel, zoals van de koe)' in value or "Une crème d'origine animale (ex: bovine, caprine, ...)" in value:
        return 1
    return 0

FL.insert(46, 'Origine_2', FL['Q170'].apply(lambda x: check_condition(x)))
WA.insert(46, 'Origine_2', WA['Q170'].apply(lambda x: check_condition(x)))

""" Origine_3: None """

def check_condition(value):
    if pd.isna(value):
        return 0
    if  'Ik heb me bij het maken van de keuzes geen specifieke herkomst van het product voorgesteld' in value or "Je n'avais pas une origine précise en tête" in value:
        return 1
    return 0

FL.insert(47, 'Origine_3', FL['Q170'].apply(lambda x: check_condition(x)))
WA.insert(47, 'Origine_3', WA['Q170'].apply(lambda x: check_condition(x)))

""" Checking frequencies """

combined_df = pd.concat([FL, WA], ignore_index=True)
print(combined_df['Origine_1'].value_counts(normalize=True) * 100)  
print(combined_df['Origine_2'].value_counts(normalize=True) * 100)  
print(combined_df['Origine_3'].value_counts(normalize=True) * 100)  

""" Deleting columns """

FL.drop(columns=['Q170'], inplace=True)
WA.drop(columns=['Q170'], inplace=True)

""" Adding the BBdiscard columns to DCE """

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Origine_1'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Origine_1'].values[0]

    return x

DCE['Origine_1'] = DCE.apply(determine, axis=1)

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Origine_2'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Origine_2'].values[0]

    return x

DCE['Origine_2'] = DCE.apply(determine, axis=1)

#%% Understanding colhumn Vraag 9: how often one smells, looks and tastes before discarding

""" Renaming the column """

FL.rename(columns={'Vraag 9 ': 'LST'}, inplace=True)
WA.rename(columns={'Vraag 9 ': 'LST'}, inplace=True)

""" Uniform values """

value_mapping = {
    'Altijd': 4,
    'Jamais': 1,
    'Nooit': 1,
    'Parfois': 2,
    'Soms': 2,
    'Souvent': 3,
    'Vaak': 3,
    'Toujours': 4
}

FL['LST'] = FL['LST'].replace(value_mapping)
WA['LST'] = WA['LST'].replace(value_mapping)

FL['LST'] = FL['LST'].fillna(0)
WA['LST'] = WA['LST'].fillna(0)

""" Checking frequencies """

combined_df = pd.concat([FL, WA], ignore_index=True)
freq = combined_df['LST'].value_counts(normalize=True) * 100  
print("Relative frequencies of unique values in the 'LST' column:")
print(freq)

""" Adding LST column to DCE """

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'LST'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'LST'].values[0]

    return x

DCE['LST'] = DCE.apply(determine, axis=1)

#%% Understanding the column Vraag 18a  - Whether there are irrelevant factors 

""" Irr """

def check_condition(value):
    if pd.isna(value):
        return 0
    if 'Ja, er waren factoren waar ik geen rekening mee heb gehouden.' in value or "Oui, il y a des facteurs que je n'ai pas pris en compte" in value:
        return 1
    return 0

FL.insert(49, 'Irr', FL['Vraag 18a '].apply(lambda x: check_condition(x)))
WA.insert(49, 'Irr', WA['Vraag 18a '].apply(lambda x: check_condition(x)))

""" Checking frequencies """

combined_df = pd.concat([FL, WA], ignore_index=True)
print(combined_df['Irr'].value_counts(normalize=True) * 100)  

""" Dropping the column """

FL.drop(columns=['Vraag 18a '], inplace=True)
WA.drop(columns=['Vraag 18a '], inplace=True)

#%% Understanding the column Vraag 18b  - Which irrelevant factors 

""" Irr_attr_1: whether the package was opened or not """

def check_condition(value):
    if pd.isna(value):
        return 0
    if '(On)geopend' in value or "La date d'ouverture du paquet" in value:
        return 1
    return 0

FL.insert(50, 'Irr_attr_1', FL['Vraag 18b'].apply(lambda x: check_condition(x)))
WA.insert(50, 'Irr_attr_1', WA['Vraag 18b'].apply(lambda x: check_condition(x)))


""" Irr_attr_2: type of packaging """

def check_condition(value):
    if pd.isna(value):
        return 0
    if 'Verpakking' in value or "Le type d'emballage" in value:
        return 1
    return 0

FL.insert(51, 'Irr_attr_2', FL['Vraag 18b'].apply(lambda x: check_condition(x)))
WA.insert(51, 'Irr_attr_2', WA['Vraag 18b'].apply(lambda x: check_condition(x)))

""" Irr_attr_3: best before date """

def check_condition(value):
    if pd.isna(value):
        return 0
    if 'THT datum' in value or "Date de durabilité minimale (DDM)" in value:
        return 1
    return 0

FL.insert(52, 'Irr_attr_3', FL['Vraag 18b'].apply(lambda x: check_condition(x)))
WA.insert(52, 'Irr_attr_3', WA['Vraag 18b'].apply(lambda x: check_condition(x)))

""" Irr_attr_4: informational label """

def check_condition(value):
    if pd.isna(value):
        return 0
    if 'Soort label' in value or "Label de contrôle" in value:
        return 1
    return 0

FL.insert(53, 'Irr_attr_4', FL['Vraag 18b'].apply(lambda x: check_condition(x)))
WA.insert(53, 'Irr_attr_4', WA['Vraag 18b'].apply(lambda x: check_condition(x)))

""" Checking frequencies """

combined_df = pd.concat([FL, WA], ignore_index=True)
print(combined_df['Irr_attr_1'].value_counts(normalize=True) * 100)  
print(combined_df['Irr_attr_2'].value_counts(normalize=True) * 100)  
print(combined_df['Irr_attr_3'].value_counts(normalize=True) * 100)  
print(combined_df['Irr_attr_4'].value_counts(normalize=True) * 100)  


""" Dropping the column """

FL.drop(columns=['Vraag 18b'], inplace=True)
WA.drop(columns=['Vraag 18b'], inplace=True)

""" Adding Irr_attr columns to DCE """

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Irr_attr_1'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Irr_attr_1'].values[0]

    return x

DCE['Irr_attr_1'] = DCE.apply(determine, axis=1)

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Irr_attr_2'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Irr_attr_2'].values[0]

    return x

DCE['Irr_attr_2'] = DCE.apply(determine, axis=1)

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Irr_attr_3'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Irr_attr_3'].values[0]

    return x

DCE['Irr_attr_3'] = DCE.apply(determine, axis=1)

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Irr_attr_4'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Irr_attr_4'].values[0]

    return x

DCE['Irr_attr_4'] = DCE.apply(determine, axis=1)

#%% Understanding the column Q171  - Any missing attributes?

""" Renaming the column """

FL.rename(columns={'Q171': 'Missing'}, inplace=True)
WA.rename(columns={'Q171': 'Missing'}, inplace=True)

""" Uniform values """

value_mapping = {
    'Neen, de belangrijkste factoren waren aanwezig in het experiment.': 'No',
    'Ja, namelijk:': 'Yes',
    "Non, les facteurs les plus importants étaient présents dans l’expérience": 'No',
    'Oui, lesquels:': 'Yes'
}

FL['Missing'] = FL['Missing'].replace(value_mapping)
WA['Missing'] = WA['Missing'].replace(value_mapping)

FL['Missing'] = FL['Missing'].fillna('Unknown')
WA['Missing'] = WA['Missing'].fillna('Unknown')

""" Checking frequencies """

combined_df = pd.concat([FL, WA], ignore_index=True)
freq = combined_df['Missing'].value_counts(normalize=True) * 100  
print("Relative frequencies of unique values in the 'Missing' column:")
print(freq)


#%% Which missing attributes

# Quantity of the product left
# The smell
# The taste: sour,...
# The look: color, texture
# Animal or plant based
# Nutriscore
# nog langer geleden geopend dan 7 dagen
# State of the packaging ex; convec
# whether its for them or others
# price
# label of once opened then X days good
# percentage of fat
# brand
# whether it can be closed

""" Dropping the column """

FL.drop(columns=['Q171_1_TEXT'], inplace=True)
WA.drop(columns=['Q171_1_TEXT'], inplace=True)

#%% Understanding the column Vraag 19a  - Consider any factors when shopping?

""" Renaming the column """

FL.rename(columns={'Vraag 19a ': 'Consider'}, inplace=True)
WA.rename(columns={'Vraag 19a ': 'Consider'}, inplace=True)

""" Uniform values """

value_mapping = {
    'Ja': 0,
    'Nee': 1,
    'Oui': 0,
    "Non": 1
}

FL['Consider'] = FL['Consider'].replace(value_mapping)
WA['Consider'] = WA['Consider'].replace(value_mapping)

FL['Consider'] = FL['Consider'].fillna('Unknown')
WA['Consider'] = WA['Consider'].fillna('Unknown')

""" Checking frequencies """

combined_df = pd.concat([FL, WA], ignore_index=True)
freq = combined_df['Consider'].value_counts(normalize=True) * 100  
print("Relative frequencies of unique values in the 'Consider' column:")
print(freq)

value_mapping = {
    'Unknown': 0,
}

FL['Consider'] = FL['Consider'].replace(value_mapping)
WA['Consider'] = WA['Consider'].replace(value_mapping)

#%% Understanding the column Vraag 19b  - Which factors when shopping?

""" Factors_1: best before date """

def check_condition(value):
    if pd.isna(value):
        return 0
    if 'THT datum' in value or "Date de durabilité minimale (DDM)" in value:
        return 1
    return 0

FL.insert(50, 'Factors_1', FL['Vraag 19b '].apply(lambda x: check_condition(x)))
WA.insert(50, 'Factors_1', WA['Vraag 19b '].apply(lambda x: check_condition(x)))


""" Factors_2: type of label """

def check_condition(value):
    if pd.isna(value):
        return 0
    if 'Soort label' in value or "Label de contrôle" in value:
        return 1
    return 0

FL.insert(51, 'Factors_2', FL['Vraag 19b '].apply(lambda x: check_condition(x)))
WA.insert(51, 'Factors_2', WA['Vraag 19b '].apply(lambda x: check_condition(x)))

""" Factors_3: type of packaging """

def check_condition(value):
    if pd.isna(value):
        return 0
    if 'In welke verpakking het product wordt verkocht' in value or "Le type d'emballage" in value:
        return 1
    return 0

FL.insert(52, 'Factors_3', FL['Vraag 19b '].apply(lambda x: check_condition(x)))
WA.insert(52, 'Factors_3', WA['Vraag 19b '].apply(lambda x: check_condition(x)))

""" Factors_4: whether the product is plant based """

def check_condition(value):
    if pd.isna(value):
        return 0
    if 'Of het product plantaardig is' in value or "Si le produit est d'origine végétale" in value:
        return 1
    return 0

FL.insert(53, 'Factors_4', FL['Vraag 19b '].apply(lambda x: check_condition(x)))
WA.insert(53, 'Factors_4', WA['Vraag 19b '].apply(lambda x: check_condition(x)))

""" Factors_5: whether the product is biological """

def check_condition(value):
    if pd.isna(value):
        return 0
    if 'Of het product biologisch is' in value or "Si le produit est d'origine bio" in value:
        return 1
    return 0

FL.insert(53, 'Factors_5', FL['Vraag 19b '].apply(lambda x: check_condition(x)))
WA.insert(53, 'Factors_5', WA['Vraag 19b '].apply(lambda x: check_condition(x)))

""" Checking frequencies """

combined_df = pd.concat([FL, WA], ignore_index=True)
print(combined_df['Factors_1'].value_counts(normalize=True) * 100)  
print(combined_df['Factors_2'].value_counts(normalize=True) * 100)  
print(combined_df['Factors_3'].value_counts(normalize=True) * 100)  
print(combined_df['Factors_4'].value_counts(normalize=True) * 100)  
print(combined_df['Factors_5'].value_counts(normalize=True) * 100)  

""" Dropping the column """

FL.drop(columns=['Vraag 19b '], inplace=True)
WA.drop(columns=['Vraag 19b '], inplace=True)

""" Unique answers """

# FL
# Melk, ei, vis e' vlees ben ik voorzichtig mee als deze over data zijn
# OORSPRONG VAN HET PRODUCT BELGISCHE PRODUCTEN HEBBEN VOORRANG
# Prijs
# Prijs
# bol staan, eventuele schimmel aan de buitenkant
# l'hygiène de l'étagère présentoir
# open of niet open en hoe lang
# prijs 
# temps d'ouverture

# WA
# Goût 
# L'aspect
# L'etat de l'embalage
# La traduction en français est lamentable
# REFRIGERE ou PAS

FL.drop(columns=['Vraag 19b _5_TEXT'], inplace=True)
WA.drop(columns=['Vraag 19b _5_TEXT'], inplace=True)

""" Adding the columns to DCE """

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Factors_1'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Factors_1'].values[0]

    return x

DCE['Factors_1'] = DCE.apply(determine, axis=1)

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Factors_2'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Factors_2'].values[0]

    return x

DCE['Factors_2'] = DCE.apply(determine, axis=1)

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Factors_3'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Factors_3'].values[0]

    return x

DCE['Factors_3'] = DCE.apply(determine, axis=1)

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Factors_4'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Factors_4'].values[0]

    return x

DCE['Factors_4'] = DCE.apply(determine, axis=1)

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Factors_5'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Factors_5'].values[0]

    return x

DCE['Factors_5'] = DCE.apply(determine, axis=1)

#%% Understanding the column Vraag 20a  - Whether they know the difference in date labels

""" Renaming the column """

FL.rename(columns={'Vraag 20a ': 'Diff'}, inplace=True)
WA.rename(columns={'Vraag 20a ': 'Diff'}, inplace=True)

""" Uniform values """

value_mapping = {
    'Ja, ik was me bewust van het verschil met de "Te Gebruiken Tot (TGT)" datum': 1,
    'Nee, ik beschouwde de THT als de "Te Gebruiken Tot (TGT)" datum': 0,
    "Oui, j'étais conscient(e) de la différence avec la date limite de consommation (DLC)": 1,
    "Non, pour moi la DDM et la DLC étaient identiques": 0
}

FL['Diff'] = FL['Diff'].replace(value_mapping)
WA['Diff'] = WA['Diff'].replace(value_mapping)

FL['Diff'] = FL['Diff'].fillna('Unknown')
WA['Diff'] = WA['Diff'].fillna('Unknown')

""" Checking frequencies """

combined_df = pd.concat([FL, WA], ignore_index=True)
freq = combined_df['Diff'].value_counts(normalize=True) * 100  
print("Relative frequencies of unique values in the 'Diff' column:")
print(freq)

value_mapping = {
    'Unknown': 0,
}

FL['Diff'] = FL['Diff'].replace(value_mapping)
WA['Diff'] = WA['Diff'].replace(value_mapping)

""" Adding it to the DCE """

def determine(row):
    if row['Region'] == 1:
        x = FL.loc[FL['ID'] == row['ID'], 'Diff'].values[0]
    else:
        x = WA.loc[WA['ID'] == row['ID'], 'Diff'].values[0]

    return x

DCE['Diff'] = DCE.apply(determine, axis=1)

#%% Adding a Gender column to DCE

DCE['Gender'] = 0

for index, row in DCE.iterrows():
    if row['Region'] == 1:
        fl_row = FL[FL['ID'] == row['ID']]
        if not fl_row.empty and fl_row['Gender'].iloc[0] == 1:
            DCE.at[index, 'Gender'] = 1
    elif row['Region'] == 0:
        wa_row = WA[WA['ID'] == row['ID']]
        if not wa_row.empty and wa_row['Gender'].iloc[0] == 1:
            DCE.at[index, 'Gender'] = 1
            
del fl_row, index, row, wa_row

#%% Adding Education to DCE

DCE['Education'] = 0

# Iterate over each row of DCE to check and update 'Education' column
for index, row in DCE.iterrows():
    if row['Region'] == 1:
        fl_row = FL[FL['ID'] == row['ID']]
        if not fl_row.empty:
            education = fl_row['Education'].iloc[0]
            if education == 'Primary Education':
                DCE.at[index, 'Education'] = 1
            elif education == 'Secondary Education':
                DCE.at[index, 'Education'] = 2
            elif education == 'Bachelor':
                DCE.at[index, 'Education'] = 3
            elif education == 'Master':
                DCE.at[index, 'Education'] = 4
            elif education == 'PhD':
                DCE.at[index, 'Education'] = 5
            else:
                DCE.at[index, 'Education'] = 0
    elif row['Region'] == 0:
        wa_row = WA[WA['ID'] == row['ID']]
        if not wa_row.empty:
            education = wa_row['Education'].iloc[0]
            if education == 'Primary Education':
                DCE.at[index, 'Education'] = 1
            elif education == 'Secondary Education':
                DCE.at[index, 'Education'] = 2
            elif education == 'Bachelor':
                DCE.at[index, 'Education'] = 3
            elif education == 'Master':
                DCE.at[index, 'Education'] = 4
            elif education == 'PhD':
                DCE.at[index, 'Education'] = 5
            else:
                DCE.at[index, 'Education'] = 0

del education, fl_row, index, row, wa_row

#%% Adding the Household variable to DCE

DCE['Household tally'] = 0

# Iterate over each row of DCE to check and update 'Household Tally' column
for index, row in DCE.iterrows():
    if row['Region'] == 1:
        fl_row = FL[FL['ID'] == row['ID']]
        if not fl_row.empty:
            household = fl_row['Household tally'].iloc[0]
            if pd.notna(household):
                DCE.at[index, 'Household tally'] = household
            else:
                DCE.at[index, 'Household tally'] = 1
    elif row['Region'] == 0:
        wa_row = WA[WA['ID'] == row['ID']]
        if not wa_row.empty:
            household = wa_row['Household tally'].iloc[0]
            if pd.notna(household):
                DCE.at[index, 'Household tally'] = household
            else:
                DCE.at[index, 'Household tally'] = 1

del fl_row, index, row, wa_row
del household

#%% Adding the income variable 

""" Coding the variable income as continuous """

mapping = {
    '€1001 – €2000': 1,
    "€2001 – €3000": 2,
    '€3001 – €4000': 3,
    '€4001 – €5000': 4,
    '> €5000': 5
}

WA['Income per month'] = WA['Income per month'].replace(mapping)
FL['Income per month'] = FL['Income per month'].replace(mapping)

""" Adding the variable to DCE dataset """

DCE['Income per month'] = 0

for index, row in DCE.iterrows():
    if row['Region'] == 1:
        fl_row = FL[FL['ID'] == row['ID']]
        if not fl_row.empty:
            income = fl_row['Income per month'].iloc[0]
            if pd.notna(income):
                DCE.at[index, 'Income per month'] = income
            else:
                DCE.at[index, 'Income per month'] = 0
    elif row['Region'] == 0:
        wa_row = WA[WA['ID'] == row['ID']]
        if not wa_row.empty:
            income = wa_row['Income per month'].iloc[0]
            if pd.notna(income):
                DCE.at[index, 'Income per month'] = income
            else:
                DCE.at[index, 'Income per month'] = 0

del fl_row, index, row, wa_row, mapping

#%% Adding the Age variable 

""" Coding the variable income as continuous """

mapping = {
    '18-24': 1,
    "25-34": 2,
    '35-44': 3,
    '45-54': 4,
    '55-64': 5,
    '65+': 6
}

WA['Age'] = WA['Age'].replace(mapping)
FL['Age'] = FL['Age'].replace(mapping)

""" Adding the variable to DCE dataset """

DCE['Age'] = 0

for index, row in DCE.iterrows():
    if row['Region'] == 1:
        fl_row = FL[FL['ID'] == row['ID']]
        if not fl_row.empty:
            Age = fl_row['Age'].iloc[0]
            if pd.notna(Age):
                DCE.at[index, 'Age'] = Age
            else:
                DCE.at[index, 'Age'] = 0
    elif row['Region'] == 0:
        wa_row = WA[WA['ID'] == row['ID']]
        if not wa_row.empty:
            Age = wa_row['Age'].iloc[0]
            if pd.notna(Age):
                DCE.at[index, 'Age'] = Age
            else:
                DCE.at[index, 'Age'] = 0

del fl_row, index, row, wa_row, mapping

#%% Adding the Years in Belgium variable 

""" Coding the variable income as continuous """

mapping = {
    'Always': 5,
    "> 10": 4,
    '5 or less': 2,
    '10 or less': 3,
    '1 or less': 1

}

WA['Years in Belgium'] = WA['Years in Belgium'].replace(mapping)
FL['Years in Belgium'] = FL['Years in Belgium'].replace(mapping)

""" Adding the variable to DCE dataset """

DCE['Years in Belgium'] = 0

for index, row in DCE.iterrows():
    if row['Region'] == 1:
        fl_row = FL[FL['ID'] == row['ID']]
        if not fl_row.empty:
            years = fl_row['Years in Belgium'].iloc[0]
            if pd.notna(years):
                DCE.at[index, 'Years in Belgium'] = years
            else:
                DCE.at[index, 'Years in Belgium'] = 0
    elif row['Region'] == 0:
        wa_row = WA[WA['ID'] == row['ID']]
        if not wa_row.empty:
            years = wa_row['Years in Belgium'].iloc[0]
            if pd.notna(years):
                DCE.at[index, 'Years in Belgium'] = years
            else:
                DCE.at[index, 'ears in Belgium'] = 0

del fl_row, index, row, wa_row, mapping
                
#%% Renaming the column Sets to Blocks

DCE.rename(columns={'Sets': 'Blocks'}, inplace=True)

#%% Add a new ID column that Stata can read called Respondent

respondent_sequence = [i for i in range(1, len(DCE) // 12 + 2) for _ in range(12)]
DCE.insert(0, 'Respondent', respondent_sequence[:len(DCE)])

del respondent_sequence


#%% Saving the DCE data as a xlsx

file_path = "DCE_data.xlsx"
DCE.to_excel(file_path, index=False)

""" Adjusting the coloring of the excel sheet """

wb = load_workbook(file_path)
ws = wb.active
light_blue_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
dark_blue_fill = PatternFill(start_color="8DB6CD", end_color="8DB6CD", fill_type="solid")  # Slightly darker than light blue

for col in range(1, ws.max_column + 1):
    ws.cell(row=1, column=col).fill = dark_blue_fill

for row in range(2, ws.max_row + 1):
    if (row - 2) % 24 < 12:
        fill_color = light_blue_fill
    else:
        fill_color = dark_blue_fill
    for col in range(1, ws.max_column + 1):
        ws.cell(row=row, column=col).fill = fill_color

wb.save(file_path)

del light_blue_fill, row, wb, ws, fill_color, dark_blue_fill, col

file_path = "/Users/thijshanssen/Desktop/THESIS/Data/DCE_data.xlsx"
df = pd.read_excel(file_path)
csv_file_path = "DCE_data_modified.csv"
df.to_csv(csv_file_path, index=False)



